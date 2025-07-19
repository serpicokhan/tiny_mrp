from django.shortcuts import render,redirect, get_object_or_404
from django.http import JsonResponse,HttpResponse
from django.template.loader import render_to_string
from django.views.decorators.csrf import csrf_exempt
from mrp.models import PurchaseRequest, RequestItem,SysUser,Part,Asset2,PurchaseRequestFile,PurchaseRequestFaktor,Asset,Comment,PurchaseNotes,PurchaseActivityLog,RFQ
from django.template.loader import render_to_string
from mrp.business.purchaseutility import *
from mrp.business.DateJob import *
from django.db.models import Q
from django.contrib.auth.context_processors import PermWrapper
from django.core.files.storage import FileSystemStorage
from mrp.forms import PurchaseRequestFileForm,RFQForm
import openpyxl
from openpyxl.styles import Border, Side, PatternFill,Font,Alignment
import json
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import permission_required
import requests as rqt
from django.contrib.auth.models import User
from django.db.models import Max
import ast
from django.contrib.auth.models import User, Permission
import re
from mrp.forms import RequestItemForm

@login_required

def list_purchase(request):
    return render(request,"mrp/purchase/purchase.html",{})
@login_required

def list_purchase_req(request):
    search_query = request.GET.get('q', '').strip() 
    start = request.GET.get('start', False) 
    end = request.GET.get('end', False)

    sort_by = request.GET.get('sort_by', '-id')  # Default sorting by `created_at` in descending order
    status_filter = request.GET.get('status', 'all')  # Default to show all statuses
    requests=PurchaseRequest.objects.filter(user__userId=request.user).order_by('-id')
    if search_query:
        filters = Q(items__item_name__partName__icontains=search_query) | \
                Q(items__description__icontains=search_query) | \
                Q(user__fullName__icontains=search_query)
        
        # Only add the id filter if the search query is a digit
        if search_query.isdigit():
            filters |= Q(id=search_query)
        
        requests = requests.filter(filters).distinct()
    # Status filtering
    if status_filter != 'all':
        requests = requests.filter(status=status_filter)
    valid_sort_fields = ['id', '-id', 'created_at', '-created_at', 'status', '-status']
    if sort_by in valid_sort_fields:
        requests = requests.order_by(sort_by)
    if start and end:
        # print(start,end,'!!!!!!!!!!!!!!!!!')
        start_of_month=DateJob.getTaskDate(start)
        end_of_month=DateJob.getTaskDate(end)
        # print(start,end,'!!!!!!!!!!!!!!!!!')

        requests=requests.filter(created_at__range=[start_of_month,end_of_month])
        start_of_month=start_of_month.strftime('%Y-%m-%d')
        end_of_month=end_of_month.strftime('%Y-%m-%d')
    else:
        today_shamsi = jdatetime.date.today()
        start_of_month = jdatetime.date(today_shamsi.year, today_shamsi.month, 1)

        # Calculate the end of the current month
        if today_shamsi.month < 12:  # Not the last month of the year
            next_month = jdatetime.date(today_shamsi.year, today_shamsi.month + 1, 1)
        else:  # For Esfand (last month), move to Farvardin of the next year
            next_month = jdatetime.date(today_shamsi.year + 1, 1, 1)

        end_of_month = next_month - jdatetime.timedelta(days=1)
        start_of_month=start_of_month.togregorian().strftime('%Y-%m-%d')
        end_of_month=end_of_month.togregorian().strftime('%Y-%m-%d')
    
    
    ws=PurchaseUtility.doPaging(request,requests)
    return render(request,"mrp/purchase/purchaseList.html",
                  {
                    "req":ws,
                    "search_query": search_query,
                    "sort_by": sort_by,
                    "status": status_filter,
                    'perms': PermWrapper(request.user),
                    'users':SysUser.objects.all(),
                    'start':start_of_month,
                    'end':end_of_month,
                    
                    })
@login_required
@permission_required('mrp.view_all_request',login_url='/Purchases')

def list_purchase_req_detail(request):
    search_query = request.GET.get('q', '').strip() 
    start = request.GET.get('start', False) 
    end = request.GET.get('end', False)
    userlist = request.GET.getlist('userlist')
    userlist = [int(user_id) for user_id in userlist if user_id.isdigit()]
    
    is_tamiri = request.GET.get('customSwitch3_',False)
    
    

    sort_by = request.GET.get('sort_by', '-id')  # Default sorting by `created_at` in descending order
    status_filter = request.GET.get('status', 'all')  # Default to show all statuses
    
    requests=PurchaseRequest.objects.all().order_by('-created_at')

    if search_query:
        filters = Q(items__item_name__partName__icontains=search_query) | \
                Q(items__description__icontains=search_query) | \
                Q(user__fullName__icontains=search_query)
        
        # Only add the id filter if the search query is a digit
        if search_query.isdigit():
            filters |= Q(id=search_query)
        
        requests = requests.filter(filters).distinct()
    # Status filtering
    if status_filter != 'all':
        requests = requests.filter(status=status_filter)

    valid_sort_fields = ['id', '-id', 'created_at', '-created_at', 'status', '-status']
    if sort_by == 'latest_activity_timestamp':
    # Annotate PurchaseRequest with the latest activity log timestamp
        requests = requests.annotate(
            latest_activity_timestamp=Max('plogs__timestamp')
        ).order_by('-latest_activity_timestamp')
    elif sort_by in valid_sort_fields:
    # If the sort_by is a valid field, apply the sorting
        requests = requests.order_by(sort_by)
    


        
    if start and end:
        print(start,end,'!!!!!!!!!!!!!!!!!')
        start_of_month=DateJob.getTaskDate(start)
        end_of_month=DateJob.getTaskDate(end)
        # print(start,end,'!!!!!!!!!!!!!!!!!')

        requests=requests.filter(created_at__range=[start_of_month,end_of_month])
        start_of_month=start_of_month.strftime('%Y-%m-%d')
        end_of_month=end_of_month.strftime('%Y-%m-%d')
    else:
        today_shamsi = jdatetime.date.today()
        start_of_month = jdatetime.date(today_shamsi.year, today_shamsi.month, 1)

        # Calculate the end of the current month
        if today_shamsi.month < 12:  # Not the last month of the year
            next_month = jdatetime.date(today_shamsi.year, today_shamsi.month + 1, 1)
        else:  # For Esfand (last month), move to Farvardin of the next year
            next_month = jdatetime.date(today_shamsi.year + 1, 1, 1)

        end_of_month = next_month - jdatetime.timedelta(days=1)
        start_of_month=start_of_month.togregorian().strftime('%Y-%m-%d')
        end_of_month=end_of_month.togregorian().strftime('%Y-%m-%d')
    
    if(userlist):
        requests=requests.filter(user__id__in=userlist)

    
    if(is_tamiri=="on"):
        requests=requests.filter(is_tamiri="True")
    ws=PurchaseUtility.doPaging(request,requests)
    
    return render(request,"mrp/purchase/purchaseList2.html",
                  {
                    "req":ws,
                    "search_query": search_query,
                    "sort_by": sort_by,
                    "status": status_filter,
                    'perms': PermWrapper(request.user),
                    'users':SysUser.objects.all(),
                    'start':start_of_month,
                    'end':end_of_month,
                    'userlist':userlist,
                    'is_tamiri':is_tamiri,
                    'start_j':start,
                    'end_j':end
                    })
@csrf_exempt
def save_purchase_request(request):
    if request.method == 'POST':
        # try:
            data = json.loads(request.body)
            items = data.get('items', [])
            req_id=data.get('id', False)
            created_at=data.get("created_at",False)
            is_emergency=data.get("emergency",False)
            is_tamiri=data.get("tamiri",False)
            print(is_tamiri,"tamiri")


            
            purchase_request=None
            # Create a new PurchaseRequest
            r_user=data.get('user_name', False)
            if(r_user):
                r_user=SysUser.objects.get(userId=r_user)
            else:
                r_user=SysUser.objects.get(userId=request.user)            
            if(req_id):
                #update
                purchase_request = get_object_or_404(PurchaseRequest, id=req_id)
                # purchase_request.created_at=DateJob.getTaskDate(created_at)
                purchase_request.save()
                existing_item_ids = [item.get('id') for item in items if 'id' in item]
                RequestItem.objects.filter(purchase_request=purchase_request).exclude(id__in=existing_item_ids).delete()
                for item in items:
                    if ('id' in item) and (item['id']):  # Update existing item
                        request_item = get_object_or_404(RequestItem, id=item['id'], purchase_request=purchase_request)
                        request_item.item_name = Part.objects.get(id=item['part_code'])
                        request_item.quantity = item['quantity']
                        request_item.consume_place = Asset2.objects.get(id=item['machine_code'])
                        request_item.description = item['description']
                        request_item.save()
                    else:  # Create new item
                        RequestItem.objects.create(
                            purchase_request=purchase_request,
                            item_name=Part.objects.get(id=item['part_code']),
                            quantity=item['quantity'],
                            consume_place=Asset2.objects.get(id=item['machine_code']),
                            description=item['description']
                        )
                # print(existing_item_ids)
            else:
                
                purchase_request = PurchaseRequest.objects.create(
                user=r_user,
                is_emergency=is_emergency,  # Assuming user is logged in
                is_tamiri=is_tamiri,  # Assuming user is logged in
                created_at=DateJob.getTaskDate(created_at)
                # consume_place="General",  # Default or get from frontend if applicable
                # description="Auto-generated request"
                )   
                PurchaseActivityLog.objects.create(
                    user=request.user.sysuser,  # User making the change
                    purchase_request=purchase_request,
                    action=f"{request.user.sysuser} درخواست را ایجاد نمود"
                )



            # Add items to the PurchaseRequest
                for item in items:
                    RequestItem.objects.create(
                        purchase_request=purchase_request,
                        item_name=Part.objects.get(id=item['part_code']),
                        quantity=item['quantity'],
                        consume_place=Asset2.objects.get(id=item['machine_code']),
                        description=item['description']

                    )
            # list_item=list_purchaseRequeset(request)
            data=dict()
            data["parchase_req_html"]=render_to_string('mrp/purchase/partialPurchaseList.html', {
                        
                        
                        'perms': PermWrapper(request.user) 

                       

                        
                    },request)
            data["http_status"]="ok"
            data['purchase_request']=purchase_request.id

            return JsonResponse(data)

        # except Exception as e:
        #     print('!!!!!!',e)
        #     return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)

def create_purchase(request):
    if(request.method=="GET"):
        data=dict()
        sysusers=SysUser.objects.all()
        data["parchase_req_html"]=render_to_string('mrp/purchase/createReq.html', {
                'users': sysusers,

                
            },request)
        return JsonResponse(data)
def update_purchase(request,id):
    company=PurchaseRequest.objects.get(id=id)
    req_items=RequestItem.objects.filter(purchase_request=company)
    comments = company.comments.all() 
    notes = company.notes.filter(user=request.user.sysuser) 
    files=PurchaseRequestFile.objects.filter(purchase_request=company)
    if(request.method=="GET"):
        data=dict()
        data["parchase_req_html"]=render_to_string('mrp/purchase/updateReq.html', {
                'company': company,
                'items':req_items,
                'files':files,
                'date_':company.created_at.strftime('%Y/%m/%d'),
                "comments": comments,
                'perms': PermWrapper(request.user) 

                
                
            },request)
        data["parchase_req_tab"]=render_to_string('mrp/purchase/tab_updatereq.html', {
                'company': company,
        
                'files':files,
                'date_':company.created_at.strftime('%Y/%m/%d'),
                "comments": comments,
                'notes':notes


                
            },request)
        return JsonResponse(data)

def update_purchase_v2(request,id):
    print("here!!!!",'!!!!!!!!!!!!!!')
    company=PurchaseRequest.objects.get(id=id)
    req_items=RequestItem.objects.filter(purchase_request=company)
    files=PurchaseRequestFile.objects.filter(purchase_request=company)
    faktors=PurchaseRequestFaktor.objects.filter(purchase_request=company)
    comments = company.comments.all().filter(Q(to_user=request.user.sysuser) | Q(to_user__isnull=True)| Q(user__userId=request.user))

    # notes = company.notes.filter(user=request.user.sysuser) 
    notes=None
    user_groups = request.user.groups.all()
    if user_groups.exists():
        notes = company.notes.filter(user__userId__groups__in=user_groups)
    else:
        notes = company.notes.filter(user=request.user.sysuser)
    plogs=company.plogs.all()
    rfqs=RFQ.objects.filter(items__purchase_request=company)



    if(request.method=="GET"):
        data=dict()
        data["parchase_req_html"]=render_to_string('mrp/purchase/updateReq_v2.html', {
                'company': company,
                'items':req_items,
                'files':files,
                'date_':company.created_at.strftime('%Y/%m/%d'),
                "comments": comments,
                "notes":notes,
                "logs":plogs,
                "faktors":faktors,
                "rfqs":rfqs,

                'perms': PermWrapper(request.user) 


                
            },request)
        data["parchase_req_tab"]=render_to_string('mrp/purchase/tab_updatereq_v2.html', {
                'company': company,
        
                'files':files,
                'date_':company.created_at.strftime('%Y/%m/%d'),
                "comments": comments,
                "notes":notes,
                "logs":plogs,
               



                
            },request)
        return JsonResponse(data)
def user_input_list():
    try:
        # Replace 'myapp.some_permission' with your specific permission codename
        permission = Permission.objects.get(codename='can_operator_purchase', content_type__app_label='myapp')
        users = User.objects.filter(user_permissions=permission)
        sysusers=SysUser.objects.filter(userId__in=users)
        dict_list = [{'key': user.fullName, 'value': user.id} for user in sysusers]
    except Permission.DoesNotExist:
        dict_list = []
    return dict_list()
def purchase_dash(request):
    list_items=PurchaseRequest.objects.filter(status='Pending').order_by('-id')
    list_items=PurchaseUtility.doPaging(request,list_items)
    
    return render(request,"mrp/purchase/mainList.html",{'items':list_items})
@csrf_exempt
def confirm_request(request,id):
    company = PurchaseRequest.objects.get(id=id)

    # Define the group-to-status mapping
    group_status_map = {
        "anbar": ["Approved"],  # انبار
        "managers": ["Approve2"],          # مدیر
        "director": ["Approve3"],          # مدیرعامل
        "purchase": ["Approve4","Ordered","Purchased"],         # خرید
    }

    # Define the status hierarchy
    status_hierarchy = [
        "Pending", "Approved", "Approve2","Approve4", "Approve3", "Ordered","Purchased"
    ]

    # Check the user's groups and determine the new status
    user_groups = request.user.groups.values_list('name', flat=True)
    new_status = None
    next_status=None
    next_group=None

    for group_name, statuses in group_status_map.items():
        if group_name in user_groups:
            

            # Filter the group's statuses to find the next status in sequence
            sorted_statuses = sorted(
                statuses, key=lambda status: status_hierarchy.index(status)
            )  # Sort statuses in order of the hierarchy
           
           

            for status in sorted_statuses:
                if status_hierarchy.index(status) > status_hierarchy.index(company.status):
                    new_status = status
                    # if(sorted_statuses.index(status)+1<len(sorted_statuses)):
                    #     next_status=sorted_statuses[sorted_statuses.index(status)+1]
                    #     next_group = next((group for group, statuses in group_status_map.items() if next_status in statuses), None)
                    #     next_to_next_group_users = list(User.objects.filter(groups__name=next_group))

                    break
            break
    # Find the **next-to-next** status and group
    if new_status:
        current_index = status_hierarchy.index(new_status)
        if current_index + 1 < len(status_hierarchy):
            next_status = status_hierarchy[current_index + 1]
            next_group = next((group for group, statuses in group_status_map.items() if next_status in statuses), None)

    # Find users in the next group


    # If no matching group is found, return an error
    if not new_status:
        return JsonResponse({
            "http_status": "error",
            "message": "شما نمی‌توانید این درخواست را تأیید کنید .",
        }, status=201)

    # Ensure the user cannot confirm a request if the status is already higher
    current_status_index = status_hierarchy.index(company.status)
    new_status_index = status_hierarchy.index(new_status)
    print("status",new_status,company.status)
    print("status",new_status_index,current_status_index)
    if new_status_index <= current_status_index:
        return JsonResponse({
            "http_status": "error",
            # "message": "You cannot confirm this request because the current status is already higher or equal.",
            "message": "شما نمی‌توانید این درخواست را تأیید کنید زیرا وضعیت فعلی برابر یا بالاتر از سطح موردنظر است."

        }, status=201)
    elif new_status_index > current_status_index+1:
        if(not "director" in user_groups):
            return JsonResponse({
                "http_status": "error",
                # "message": "You cannot confirm this request because the current status is already higher or equal.",
                "message": "شما نمی‌توانید این درخواست را تأیید کنید ."

            }, status=201)
        else:
            new_status="Approve3"



    # Update the status
    company.status = new_status
    company.save()
    PurchaseActivityLog.objects.create(
            user=request.user.sysuser,  # User making the change
            purchase_request=company,
            action=f"{request.user.sysuser} درخواست را تایید نمود"
        )
    ###########Send whatsapp#############

    


    # Refresh the purchase request list
    # list_item=list_purchaseRequeset(request)
    data=dict()
    data["group"]=next_group
    data["purchase_request_id"]=company.id
    # data["parchase_req_html"]=render_to_string('mrp/purchase/partialPurchaseList_v2.html', {
                
    #             'req':list_item,     
    #             'perms': PermWrapper(request.user)           
    #         },request)
    data["http_status"]="ok"
    data["status"]=company.get_status_display()


    return JsonResponse(data)
    # return JsonResponse({"status":"ok",'status':company.status})

@csrf_exempt
def reject_request(request,id):
    if(request.method=="GET"):
        print("get")
        data=dict()
        data["http_status"]="ok"
        data["parchase_req_html"]=render_to_string('mrp/purchase/partialRejectForm.html', {                    
                    'id':id,
                    'perms': PermWrapper(request.user)                     
                },request)
        return JsonResponse(data)    
    elif(request.method=="POST"):
        
        purchase_request = get_object_or_404(PurchaseRequest, id=id)        
        # Permission check
        if not purchase_request.can_be_rejected_by(request.user):
            return JsonResponse({
                'success': False, 
                'error': 'شما اجازه رد این درخواست را ندارید.'
            }, status=403)
        
        # Status check
        if purchase_request.status in ['Rejected', 'Completed','Ordered','Purchased']:
            return JsonResponse({
                'success': False, 
                'error': 'این درخواست قبلاً پردازش شده است.'
            }, status=400)
        
        try:
            
            rejection_reason = request.POST.get('rejectReason', False)

            
            if not rejection_reason:
                return JsonResponse({
                    'success': False, 
                    'error': 'دلیل رد درخواست الزامی است.'
                }, status=400)
            
            # Reject the request
            purchase_request.reject_request(
                rejected_by_user=request.user.sysuser,
                rejection_reason=rejection_reason
            )
            
            return JsonResponse({
                'success': True, 
                'message': f'درخواست خرید شماره {id} با موفقیت رد شد.',
                'new_status': purchase_request.get_status_display(),
                'rejected_by': str(purchase_request.rejected_by),
                'rejected_at': purchase_request.rejected_at.strftime('%Y-%m-%d %H:%M') if purchase_request.rejected_at else None
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False, 
                'error': 'داده های ارسالی نامعتبر است.'
            }, status=400)
        except Exception as e:
            print(e)
            return JsonResponse({
                'success': False, 
                'error': f'خطا در پردازش درخواست: {str(e)}'
            }, status=500)

def list_purchaseRequeset(request):
    return filter_request_by(request)

@csrf_exempt

def delete_purchase_request(request,id):
    company=  get_object_or_404(PurchaseRequest, id=id)
    if(request.method=="POST"):
        data=dict()
        if(request.user.is_superuser):
            PurchaseActivityLog.objects.create(
                    user=request.user.sysuser,  # User making the change
                    purchase_request=company,
                    action=f"{request.user.sysuser} درخواست را حذف نمود"
                )
            company.delete()
            list_item=list_purchaseRequeset(request)
            data["parchase_req_html"]=render_to_string('mrp/purchase/partialPurchaseList.html', {
                        
                        'req':list_item,
                      'perms': PermWrapper(request.user)                         
                    })
            data["http_status"]="ok"
            data["status"]=company.status            
        else:
            if(company.status=="Pending"):
                PurchaseActivityLog.objects.create(
                    user=request.user.sysuser,  # User making the change
                    purchase_request=company,
                    action=f"{request.user.sysuser} درخواست را حذف نمود"
                )
                company.delete()
                list_item=list_purchaseRequeset(request)
                data["parchase_req_html"]=render_to_string('mrp/purchase/partialPurchaseList.html', {
                            
                            'req':list_item,
                        'perms': PermWrapper(request.user) 


                            
                        })
                data["http_status"]="ok"
                data["status"]=company.status
            else:
                data["http_status"]="ok"
                data["status"]=company.status
                data["message"]="حدف درخواست به خاطر تغییر وضعیت امکان پذیر نمی باشد"



        return JsonResponse(data)
    return JsonResponse({'stats':'BAD!','message':'Bad Data'})
@csrf_exempt

def delete_purchase_request_v2(request,id):
    company=  get_object_or_404(PurchaseRequest, id=id)
    if(request.method=="POST"):
        data=dict()
        if(request.user.is_superuser):
            company.delete()
            list_item=list_purchaseRequeset(request)
            data["parchase_req_html"]=render_to_string('mrp/purchase/partialPurchaseList_v2.html', {
                        
                        'req':list_item,
                      'perms': PermWrapper(request.user)                         
                    })
            data["http_status"]="ok"
            data["status"]=company.status            
        else:
            if(company.status=="Pending"):
                company.delete()
                list_item=list_purchaseRequeset(request)
                data["parchase_req_html"]=render_to_string('mrp/purchase/partialPurchaseList.html', {
                            
                            'req':list_item,
                        'perms': PermWrapper(request.user) 


                            
                        })
                data["http_status"]="ok"
                data["status"]=company.status
            else:
                data["http_status"]="ok"
                data["status"]=company.status
                data["message"]="حدف درخواست به خاطر تغییر وضعیت امکان پذیر نمی باشد"



        return JsonResponse(data)
    return JsonResponse({'stats':'BAD!','message':'Bad Data'})
@csrf_exempt
def upload_purchase_images(request):
    
    print(request.method,'####################')
    if request.method == 'POST':
        purchase_request_id = request.GET.get('p_id')
        uploaded_files = request.FILES.getlist('file')
        print("files:",uploaded_files)
        print("p_id:",purchase_request_id)

        if not purchase_request_id or not uploaded_files:
            print("here!!!")


            return JsonResponse({'success': False,
                                  'errors': 'Both purchase_request and file are required.'}, status=400)

        # Validate and get the PurchaseRequest instance
        purchase_request = get_object_or_404(PurchaseRequest, id=purchase_request_id)

        # Create and save the PurchaseRequestFile instance
        for uploaded_file in uploaded_files:
            PurchaseRequestFile.objects.create(
                purchase_request=purchase_request,
                file=uploaded_file
            )

        return JsonResponse({'success': True, 'message': 'File uploaded successfully!'})
    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)
@csrf_exempt
def upload_purchase_faktors(request):
    
    print(request.method,'####################')
    if request.method == 'POST':
        purchase_request_id = request.GET.get('p_id')
        uploaded_files = request.FILES.getlist('file')
        print("files:",uploaded_files)
        print("p_id:",purchase_request_id)

        if not purchase_request_id or not uploaded_files:
            print("here!!!")


            return JsonResponse({'success': False,
                                  'errors': 'Both purchase_request and file are required.'}, status=400)

        # Validate and get the PurchaseRequest instance
        purchase_request = get_object_or_404(PurchaseRequest, id=purchase_request_id)

        # Create and save the PurchaseRequestFile instance
        for uploaded_file in uploaded_files:
            PurchaseRequestFaktor.objects.create(
                purchase_request=purchase_request,
                file=uploaded_file
            )

        return JsonResponse({'success': True, 'message': 'File uploaded successfully!'})
    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)
##########EXCEL#######################
def export_purchase_requests(request):
    # Create a workbook and a sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Purchase Requests'

    # Set row index to start at 1
    row = 1
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Define the header fill color (for PurchaseRequest and items table)
    header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    header_fill2 = PatternFill(start_color="bfb2b2", end_color="bfb2b2", fill_type="solid")
    nazanin_font = Font(name='B Nazanin', size=14)
    justified_alignment = Alignment(horizontal="center", wrap_text=True)
    

    # Get all purchase requests
    purchase_requests=filter_request_by(request)

    for purchase_request in purchase_requests:
        # Add the header for the PurchaseRequest
        sheet[f'E{row}'] = f'شماره درخواست'
        sheet[f'D{row}'] = f'کاربر'
        sheet[f'C{row}'] = f'تاریخ'
        # sheet[f'D{row}'] = f'اضطراری'
        sheet[f'B{row}'] = f"اضطراری"
        sheet[f'A{row}'] = f'وضعیت'
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = sheet[f'{col}{row}']
            cell.border = thin_border
            cell.fill = header_fill
            cell.font = nazanin_font  # Apply B Nazanin font
            # cell.alignment = Alignment(horizontal="justify")
            cell.alignment = justified_alignment 
        row += 1  # Move to the next row for the colored space
        
        sheet[f'E{row}'] = f'{purchase_request.id}'
        sheet[f'D{row}'] = f'{purchase_request.user.fullName}'
        sheet[f'C{row}'] = f'{purchase_request.get_dateCreated_jalali().strftime("%Y/%m/%d")}'
        # sheet[f'D{row}'] = f'اضطراری: {purchase_request.is_emergency}'
        sheet[f'B{row}'] = f"{'بله' if purchase_request.is_emergency else 'خیر'}"
        sheet[f'A{row}'] = f'{purchase_request.get_status_display()}'
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = sheet[f'{col}{row}']
            cell.border = thin_border
            cell.fill = header_fill
            cell.font = nazanin_font  # Apply B Nazanin font
            # cell.alignment = Alignment(horizontal="justify")
            cell.alignment = justified_alignment 
            
        
        row += 1  # Leave some space between PurchaseRequest header and items table


        # Now, create the header for the items table under each PurchaseRequest
        sheet[f'E{row}'] = 'نام کالا'
        sheet[f'D{row}'] = 'تعداد'
        sheet[f'C{row}'] = 'مورد مصرف'
        sheet[f'B{row}'] = 'شرح'
        sheet[f'A{row}'] = 'تامین کننده'
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = sheet[f'{col}{row}']
            cell.border = thin_border
            cell.fill = header_fill2
            cell.font = nazanin_font  # Apply B Nazanin font
            cell.alignment = justified_alignment 

        row += 1  # Move to the next row for the colored space
       
                

        # Fetch all items for the current purchase request
        items = purchase_request.items.all()

        for item in items:
            sheet[f'E{row}'] = item.item_name.partName  # Assuming 'partName' is the name field
            sheet[f'D{row}'] = item.quantity
            sheet[f'C{row}'] = item.consume_place.assetName  # Assuming 'name' field in Asset2
            sheet[f'B{row}'] = item.description
            sheet[f'A{row}'] = item.supplier_assigned.name if item.supplier_assigned else "مشخص نشده"
            for col in ['A', 'B', 'C', 'D', 'E']:
                cell = sheet[f'{col}{row}']
                cell.border = thin_border
                cell.font = nazanin_font  # Apply B Nazanin font
                cell.alignment = justified_alignment 

            row += 1  # Move to the next row for the next item
        
        # Leave some space between different PurchaseRequests
        row += 1  # Move to the next row for the colored space
        
     # Set the column widths
    sheet.column_dimensions['E'].width = 30  # 'Item Name' column
    sheet.column_dimensions['D'].width = 15  # 'Quantity' column
    sheet.column_dimensions['C'].width = 25  # 'Consume Place' column
    sheet.column_dimensions['B'].width = 15  # 'Price' column
    sheet.column_dimensions['A'].width = 20  # 'Supplier' column

    # Create a response to download the file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="purchase_requests_with_items.xlsx"'

    # Save the workbook to the response
    wb.save(response)
    return response


def referesh_purchase_list(request):
    search_query = request.GET.get('q', '').strip() 
    # page=request.GET.get('page', False)
    # print(page,request.GET,'@@@@@@@@@@')
    
    start = request.GET.get('start', False) 
    end = request.GET.get('end', False)
    userlist = request.GET.getlist('userlist')
    userlist = [int(user_id) for user_id in userlist if user_id.isdigit()]
    is_tamiri = request.GET.get('customSwitch3_', False)

    sort_by = request.GET.get('sort_by', '-id')  # Default sorting by `created_at` in descending order
    status_filter = request.GET.get('status', 'all')  # Default to show all statuses
    # if(request.user.is_superuser):
    if request.user.is_superuser:
        requests = PurchaseRequest.objects.all()
    else:
        # Check if the user belongs to any of the specified groups
        user_groups = request.user.groups.values_list('name', flat=True)

        # If user belongs to any of the specified groups, they can view the requests
        if any(group in user_groups for group in ['anbar', 'purchase', 'managers', 'director']):
            requests = PurchaseRequest.objects.all()  # All requests for these groups
        else:
            requests = PurchaseRequest.objects.filter(user__userId=request.user)  # Only requests for the user

    # else:
    #     requests=PurchaseRequest.objects.filter(user__userId=request.user).order_by('-created_at')

    if search_query:
        filters = Q(items__item_name__partName__icontains=search_query) | \
                Q(items__description__icontains=search_query) | \
                Q(user__fullName__icontains=search_query)
        
        # Only add the id filter if the search query is a digit
        if search_query.isdigit():
            filters |= Q(id=search_query)
        
        requests = requests.filter(filters).distinct()
    # Status filtering
    if status_filter != 'all':
        requests = requests.filter(status=status_filter)
    valid_sort_fields = ['id', '-id', 'created_at', '-created_at', 'status', '-status']
    if sort_by == 'latest_activity_timestamp':
    # Annotate PurchaseRequest with the latest activity log timestamp
        requests = requests.annotate(
            latest_activity_timestamp=Max('plogs__timestamp')
        ).order_by('-latest_activity_timestamp')
    elif sort_by in valid_sort_fields:
    # If the sort_by is a valid field, apply the sorting
        requests = requests.order_by(sort_by)

    if(userlist):
        
        requests=requests.filter(user__id__in=userlist)

    if start and end:
        print(start,end,'!!!!!!!!!!!!!!!!!')
        start_of_month=DateJob.getTaskDate(start)
        end_of_month=DateJob.getTaskDate(end)
        # print(start,end,'!!!!!!!!!!!!!!!!!')

        requests=requests.filter(created_at__range=[start_of_month,end_of_month])
        # start_of_month=start_of_month.strftime('%Y-%m-%d')
        # end_of_month=end_of_month.strftime('%Y-%m-%d')
    if(is_tamiri=="on"):
        requests=requests.filter(is_tamiri=True)
    
    
    ws= PurchaseUtility.doPaging(request,requests)
    data=dict()
    data["status"]="ok"
    data["parchase_req_html"]=render_to_string('mrp/purchase/partialPurchaseList_v2.html', {
                        
                        'req':ws,
                         'perms': PermWrapper(request.user) 

                       

                        
                    },request)
    return JsonResponse(data)



def filter_request_by(request):
    search_query = request.GET.get('q', '').strip() 
    page=request.GET.get('page', False)
    # print(page,request.GET,'@@@@@@@@@@')
    
    start = request.GET.get('start', False) 
    end = request.GET.get('end', False)
    userlist = request.GET.get('userlist', '[]')
    is_tamiri = request.GET.get('customSwitch3_', False)

    sort_by = request.GET.get('sort_by', '-id')  # Default sorting by `id` in descending order
    print(sort_by,'!!!!!!!!!!')
    status_filter = request.GET.get('status', 'all')  # Default to show all statuses
    if request.user.is_superuser:
        requests = PurchaseRequest.objects.all()
    else:
        # Check if the user belongs to any of the specified groups
        user_groups = request.user.groups.values_list('name', flat=True)

        # If user belongs to any of the specified groups, they can view the requests
        if any(group in user_groups for group in ['anbar', 'purchase', 'managers', 'director']):
            requests = PurchaseRequest.objects.all()  # All requests for these groups
        else:
            requests = PurchaseRequest.objects.filter(user__userId=request.user)  # Only requests for the user

    # if(request.user.is_superuser):
    #     requests=PurchaseRequest.objects.all().order_by('-created_at')
    # else:
    #     requests=PurchaseRequest.objects.filter(user__userId=request.user).order_by('-created_at')

    if search_query:
        filters = Q(items__item_name__partName__icontains=search_query) | \
                Q(items__description__icontains=search_query) | \
                Q(user__fullName__icontains=search_query)
        
        # Only add the id filter if the search query is a digit
        if search_query.isdigit():
            filters |= Q(id=search_query)
        
        requests = requests.filter(filters).distinct()
    # Status filtering
    if status_filter != 'all':
        requests = requests.filter(status=status_filter)
    valid_sort_fields = ['id', '-id', 'created_at', '-created_at', 'status', '-status']
    if sort_by == 'latest_activity_timestamp':
    # Annotate PurchaseRequest with the latest activity log timestamp
        requests = requests.annotate(
            latest_activity_timestamp=Max('plogs__timestamp')
        ).order_by('-latest_activity_timestamp')
    elif sort_by in valid_sort_fields:
    # If the sort_by is a valid field, apply the sorting
        requests = requests.order_by(sort_by)

    if(userlist!='[]'):
        
        # userlist2 = [int(user_id) for user_id in userlist]
        userlist2 = ast.literal_eval(userlist) 
        if isinstance(userlist2, list) and all(isinstance(i, int) for i in userlist2):
            # requests = requests.filter(user__id__in=userlist2)
            requests=requests.filter(user__id__in=userlist2)
        else:
            requests=requests.filter(user__id=userlist2)

    
    if start and end:
        print(start,end,'!!!!!!!!!!!!!!!!!')
        start_of_month=DateJob.getTaskDate(start)
        end_of_month=DateJob.getTaskDate(end)
        # print(start,end,'!!!!!!!!!!!!!!!!!')

        requests=requests.filter(created_at__range=[start_of_month,end_of_month])
        # start_of_month=start_of_month.strftime('%Y-%m-%d')
        # end_of_month=end_of_month.strftime('%Y-%m-%d')
    if is_tamiri:
        requests=requests.filter(is_tamiri=is_tamiri)
    
    
    return requests
def calendar_purchase_request_main(request):
    makan_id=request.GET.get("makan_id",False)
    makan=Asset.objects.filter(assetIsLocatedAt__isnull=True)
    return render(request,'mrp/purchase/calendar_Purchase_main.html',{'title':'تولید روزانه','makan':makan,'makan_id':int(makan_id)})
def get_purchasereq_calendar_info(request):
    # print(request.GET.get("makan"),'!!!!!!!!!!!!!!!!!!')
    print(request.GET)
    status=request.GET.get("status",'all')
    print(status)
    data=[]
    if request.user.is_superuser:
        user_info = PurchaseRequest.objects.all()
    else:
        # Check if the user belongs to any of the specified groups
        user_groups = request.user.groups.values_list('name', flat=True)

        # If user belongs to any of the specified groups, they can view the requests
        if any(group in user_groups for group in ['anbar', 'purchase', 'manager', 'director']):
            user_info = PurchaseRequest.objects.all()  # All requests for these groups
        else:
            user_info = PurchaseRequest.objects.filter(user__userId=request.user)  # Only requests for the user
    if(status!="all"):
        user_info=user_info.filter(status=status)
        # user_info = user_info.annotate(
        #         latest_activity_timestamp=Max('plogs__timestamp')
        #     ).order_by('-latest_activity_timestamp')

    # print(user_info)
    for i in user_info:
        if i.status == 'Approved':
            color = '#53c797'  # Green for approved
        elif i.status == 'Pending':
            color = '#5bc0de'  # Cyan for pending
        elif i.status == 'Rejected':
            color = '#d9534f'  # Red for rejected
        elif i.status == 'Ordered':
            color = '#5bc0de'  # Cyan for ordered
        elif i.status == 'Approve2':
            color = '#f0ad4e'  # Orange for approve2
        elif i.status == 'Approve3':
            color = '#0275d8'  # Blue for approve3
        else:
            color = '#cccccc'  # Default color if status is unknown
        last_activity_log = i.plogs.order_by('-timestamp').first()
       
        data.append({'title': f"درخواست {i.user} {i.id}",\
                'start': last_activity_log.timestamp.date() if last_activity_log else i.created_at,\
                 'color': color,\
                'id':i.id})
        # data.append({'title': f"جمع ضایعات روز: {round(z,2)}",\
        #         'start': i[0],\
        #          'color': 'red',\
        #         'id':i[0]})

    return JsonResponse(data,safe=False)
def get_purchase_request(request):
    id=request.GET.get('id',False)
    purchase=PurchaseRequest.objects.get(id=id)
    return render(request,'mrp/purchase/purchase_request_bill.html',{'purchase_request':purchase})
def add_view_by(request):
    try:
        # user_id=SysUser.objects.get(userId=request.user).id
        id=request.GET.get('id',False)
        purchase=PurchaseRequest.objects.get(id=id)
        purchase.add_viewer(request.user.id)
        return JsonResponse({},status=201)
    except Exception as e:
        print(e)
        return JsonResponse({'error':'error'},status=201)




@login_required
@csrf_exempt
def add_purchase_comment(request):
    if request.method == "POST" and request.is_ajax():
        content = request.POST.get("content")
        purchase_request_id = request.POST.get("purchase_request_id")
        parent_id = request.POST.get("parent_id")  # Optional for replies
        user = request.user.sysuser
        purchase_request = get_object_or_404(PurchaseRequest, id=purchase_request_id)
        parent_comment = Comment.objects.filter(id=parent_id).first() if parent_id else None
         # Find all @digit patterns and replace with @fullName
        def replace_user_mention(match):
            user_id = int(match.group(1))  # Extract digit after @
            user_obj = SysUser.objects.filter(id=user_id).first()
            return f"@{user_obj.fullName}" if user_obj else match.group(0)
        modified_content = re.sub(r'@(\d+)', replace_user_mention, content)
        user_id_match = re.search(r'@(\d+)', content)
        to_user = None
        if user_id_match:
            user_id = int(user_id_match.group(1))  # Extract the digit after @
            to_user = SysUser.objects.filter(id=user_id).first()  # Find SysUser with matching id
        if(user.tel1):
            url = "https://app.wallmessage.com/api/sendMessage"

            payload={
            "appkey": "78dba514-1a21-478e-8484-aecd14b198b7",
            "authkey": "ipnKtmP2bwr6t6kKDkOqV3q5w8aZcV2lLueoWBX3YlIBF1ZgMZ",
            'to': user.tel1,
            'message': f'کامنت {user.fullName} برای درخواست شماره {purchase_request_id}: {content}',
            }
            files=[]
            headers = {}
            response = rqt.request("POST", url, headers=headers, data=payload, files=files)
        comment = Comment.objects.create(
            purchase_request=purchase_request,
            user=user,
            to_user=to_user,
            content=modified_content,
            parent=parent_comment
        )

        return JsonResponse({
            "status": "success",
            "comment_id": comment.id,
            "content": modified_content,
            "created_at": comment.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "user": str(user.fullName),
            "parent_id": parent_id,
            'image':request.user.sysuser.profileImage.url
        })
    return JsonResponse({"status": "error"}, status=400)
@login_required
@csrf_exempt
def add_purchase_note(request):
    if request.method == "POST" and request.is_ajax():
        content = request.POST.get("content")
        purchase_request_id = request.POST.get("purchase_request_id")
        # parent_id = request.POST.get("parent_id")  # Optional for replies
        user = request.user.sysuser
        purchase_request = get_object_or_404(PurchaseRequest, id=purchase_request_id)
        # parent_comment = Comment.objects.filter(id=parent_id).first() if parent_id else None
        # url = "https://app.wallmessage.com/api/sendMessage"

        # payload={
        # "appkey": "7fe75ff8-b457-4abb-ad12-e4c364b79484",
        # "authkey": "06nWkgBK3SkPO1YLLC58DlxGRo7dEf3m6kV0gzsnydIgGYpfXb",
        # 'to': '09390453690',
        # 'message': f'کامنت {user.fullName} برای درخواست شماره {purchase_request_id}: {content}',
        # }
        # files=[]
        # headers = {}
        # response = rqt.request("POST", url, headers=headers, data=payload, files=files)
        comment = PurchaseNotes.objects.create(
            purchase_request=purchase_request,
            user=user,
            content=content,
           
        )

        return JsonResponse({
            "status": "success",
            "comment_id": comment.id,
            "content": comment.content,
            "created_at": comment.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "user": str(user.fullName),
            
            'image':request.user.sysuser.profileImage.url
        })
    return JsonResponse({"status": "error"}, status=400)
@csrf_exempt  # Disable CSRF for testing purposes; ensure proper CSRF handling in production
def handle_purchase_paraph(request):
    print(request.method,'##################################')
    if request.method == 'POST':
        text = request.POST.get('text')
        p_id=request.GET.get('p_id',False)
        if(p_id):
            p=PurchaseRequest.objects.get(id=p_id)
            p.manager_comment=text
            p.save()
        return JsonResponse({'message': f'Text received: {text}'})
    return JsonResponse({'error': 'Invalid request'}, status=400)
# def send_whatsapp(purchaser_req):
#      if(user.tel1):
#             url = "https://app.wallmessage.com/api/sendMessage"

#             payload={
#             "appkey": "78dba514-1a21-478e-8484-aecd14b198b7",
#             "authkey": "ipnKtmP2bwr6t6kKDkOqV3q5w8aZcV2lLueoWBX3YlIBF1ZgMZ",
#             'to': user.tel,
#             'message': f'کامنت {user.fullName} برای درخواست شماره {purchase_request_id}: {content}',
#             }
#             files=[]
#             headers = {}
#             response = rqt.request("POST", url, headers=headers, data=payload, files=files)
#############################################################
def load_more_purchaserequest(request):
    items = list_purchaseRequeset(request)  # Get all items (modify filter as needed)

    
    try:
        data = PurchaseUtility.doPaging(request,items)  # Get the requested page
    except:
        return JsonResponse({'html': ''})  # No more data, return empty response

    # Render the HTML content as a string
    html = render_to_string('mrp/purchase/partialPurchaseList_v2.html', {
                
                'req':data,     
                'perms': PermWrapper(request.user)           
            },request)

    return JsonResponse({'html': html})
def create_rfq(request,id):
    if (request.method == 'POST'):
        
        form = RFQForm(request.POST)
        # form.items=RequestItem.objects.get(id=form.items)
        return save_rfq_form(request, form, 'mrp/rfq/partialRFQCreate.html',id)
    else:

        form = RFQForm()
        return save_rfq_form(request, form, 'mrp/rfq/partialRFQCreate.html',id)
    
def save_rfq_form(request, form, template_name,id=None):

    items=RequestItem.objects.filter(purchase_request__id=id)
    data = dict()
    if (request.method == 'POST'):
        if form.is_valid():
            form.save(commit=False)
            form.instance.issued_by=request.user.sysuser
            form.save()
            data['form_is_valid'] = True
            books = RFQ.objects.filter(items__purchase_request__id=id)
            data['html_rfq_list'] = render_to_string('mrp/rfq/partialRFQList.html', {
                'rfqs': books,
                'perms': PermWrapper(request.user)
            })
        else:
            
            print(form.errors)
            data['form_is_valid'] = False

    context = {'form': form,'lid':id,'items':items}


    data['html_rfq_form'] = render_to_string(template_name, context, request=request)
    return JsonResponse(data)
##########################################################
##########################################################
def rfq_update(request, id):
    company= get_object_or_404(RFQ, id=id)
    template=""
    if (request.method == 'POST'):
        form = RFQForm(request.POST, instance=company)
    else:
        form = RFQForm(instance=company)


    return save_rfq_form(request, form,"mrp/rfq/partialRFQUpdate.html",company.items.purchase_request.id)

def rfq_delete(request, id):
    comp1 = get_object_or_404(RFQ, id=id)
    purchase_request=comp1.items.purchase_request
    data = dict()
    if (request.method == 'POST'):
        comp1.delete()
        data['form_is_valid'] = True  # This is just to play along with the existing code
        companies =  RFQ.objects.filter(items__purchase_request=purchase_request)
        #Tasks.objects.filter(maintenanceTypeId=id).update(maintenanceType=id)
        data['html_rfq_list'] = render_to_string('mrp/rfq/partialRFQList.html', {
            'rfqs': companies,
            'perms': PermWrapper(request.user)
        })
    else:
        context = {'rfq': comp1}
        data['html_rfq_form'] = render_to_string('mrp/rfq/partialRFQDelete.html',
            context,
            request=request,
        )
    return JsonResponse(data)

@csrf_exempt  # Use this decorator to exempt the view from CSRF verification for simplicity
def update_RFQ_is_verified(request, id):
    if request.method == 'POST':
        data=dict()
        user_profile = get_object_or_404(RFQ, id=id)
        user_profile.is_verified = True
        user_profile.save()
        companies =  RFQ.objects.filter(items__purchase_request=user_profile.items.purchase_request)
        #Tasks.objects.filter(maintenanceTypeId=id).update(maintenanceType=id)
        data['html_rfq_list'] = render_to_string('mrp/rfq/partialRFQList.html', {
            'rfqs': companies,
            'perms': PermWrapper(request.user)
        })
        data["is_valid"]=True
        return JsonResponse(data)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=400)
def send_purchase_wtf_msg(request):
    next_group=request.GET.get("group",False)
    id=request.GET.get("id",False)
    company=PurchaseRequest.objects.get(id=id)
    next_to_next_group_users = list(User.objects.filter(groups__name=next_group)) if next_group else []

    url = "https://app.wallmessage.com/api/sendMessage"
    if(company.status=="Approve3"):
        for next_user in next_to_next_group_users:
            

            if(next_user.sysuser.tel1):
                if(company.is_emergency):
                        payload={
                        "appkey": "78dba514-1a21-478e-8484-aecd14b198b7",
                        "authkey": "ipnKtmP2bwr6t6kKDkOqV3q5w8aZcV2lLueoWBX3YlIBF1ZgMZ",
                        'to': next_user.sysuser.tel1,
                        'message': f'⛔⛔⛔ *درخواست اضطراری* ⛔⛔⛔ \n درخواست شماره {company.id} از طرف {company.user.fullName} با مشخصات زیر نیاز به تایید شما دارد: \n\n {company.getItems3()} \n\n 【سیستم مدیریت درخواست دایانا】\n\n    ꧁ ریسندگی محتشم ꧂\n🌐 https://kth.mymrp.ir',
                        }
                else:

                    payload={
                    "appkey": "78dba514-1a21-478e-8484-aecd14b198b7",
                    "authkey": "ipnKtmP2bwr6t6kKDkOqV3q5w8aZcV2lLueoWBX3YlIBF1ZgMZ",
                    'to': next_user.sysuser.tel1,
                    'message': f'درخواست شماره {company.id} از طرف {company.user.fullName} با مشخصات زیر نیاز به تایید شما دارد: \n\n {company.getItems3()} \n\n 【سیستم مدیریت درخواست دایانا】\n\n    ꧁ ریسندگی محتشم ꧂ \n🌐 https://kth.mymrp.ir',
                    }
                
                files=PurchaseRequestFile.objects.filter(file__isnull=False,purchase_request=company)
                files2=[]
                # files=list(files)
                # for i in files:
                #     with i.file.open('rb') as file_obj:

                #         files2.append(file_obj)

                headers = {}
                response = rqt.request("POST", url, headers=headers, data=payload, files=files2)

    if(company.status=="Purchased"):
        payload={
                    "appkey": "78dba514-1a21-478e-8484-aecd14b198b7",
                    "authkey": "ipnKtmP2bwr6t6kKDkOqV3q5w8aZcV2lLueoWBX3YlIBF1ZgMZ",
                    'to': company.user.tel1,
                    'message': f'درخواست شماره {company.id} با مشخصات زیر خریداری گردید: \n\n {company.getItems3()} \n\n 【سیستم مدیریت درخواست دایانا】\n\n    ꧁ ریسندگی محتشم ꧂\n🌐 https://kth.mymrp.ir',
                    }
        files2=[]        

        headers = {}
        response = rqt.request("POST", url, headers=headers, data=payload, files=files2)
    return JsonResponse({"status":"ok"})
def purchase_request_detail_print(request, pk):
    """
    View to display the details of a specific PurchaseRequest.
    """
    # Retrieve the PurchaseRequest or return 404 if not found
    purchase_request = get_object_or_404(PurchaseRequest, pk=pk)
    
    # Optionally, add the current user to the viewed_by list
    # purchase_request.add_viewer(request.user.username)
    
    # Context to pass to the template
    context = {
        'purchase_request': purchase_request,
    }
    
    # Render the template with the context
    return render(request, 'mrp/purchase/print.html', context)
@login_required
@csrf_exempt
def update_note(request, comment_id):
    if request.method == 'POST':
        try:
            comment = PurchaseNotes.objects.get(id=comment_id)
            # Check if the user has permission to edit
            # if request.user != comment.user:
            #     return JsonResponse({
            #         'success': False,
            #         'error': 'شما اجازه ویرایش این یادداشت را ندارید'
            #     }, status=403)
            comment.content = request.POST.get('content', '')
            comment.save()
            return JsonResponse({'success': True})
        except Comment.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'یادداشت یافت نشد'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'خطا در ذخیره تغییرات: {str(e)}'
            }, status=500)
    return JsonResponse({
        'success': False,
        'error': 'درخواست نامعتبر است'
    }, status=400)

@login_required
def update_purchase_item(request,id):
    company= get_object_or_404(RequestItem, id=id)
    template=""
    if (request.method == 'POST'):
        form = RequestItemForm(request.POST, instance=company)
    else:
        form = RequestItemForm(instance=company)


    return save_purchase_item_form(request, form,"mrp/purchase/partialPurchaseItemUpdate.html")

def save_purchase_item_form(request, form, template_name):


    data = dict()
    if (request.method == 'POST'):
        if form.is_valid():
            bts=form.save()
            data['form_is_valid'] = True
            books= RequestItem.objects.filter(
    purchase_request=bts.purchase_request,
    # price=0,
    
).select_related('purchase_request').order_by('-id')
            data['html_purchase_item_list'] = render_to_string('mrp/purchase/partialPurchaseItemList.html', {
                'shifts': books,
                'perms': PermWrapper(request.user)
            })
        else:
            data['form_is_valid'] = False
            print(form.errors)

    context = {'form': form,'item_name':form.instance.item_name.partName}


    data['html_purchase_item_form'] = render_to_string(template_name, context, request=request)
    return JsonResponse(data)


@login_required
def reject_purchase_request_ajax(request, request_id):
    """
    AJAX version of rejection view for better UX
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    purchase_request = get_object_or_404(PurchaseRequest, id=request_id)
    
    # Permission check
    if not purchase_request.can_be_rejected_by(request.user):
        return JsonResponse({
            'success': False, 
            'error': 'شما اجازه رد این درخواست را ندارید.'
        }, status=403)
    
    # Status check
    if purchase_request.status in ['Rejected', 'Completed']:
        return JsonResponse({
            'success': False, 
            'error': 'این درخواست قبلاً پردازش شده است.'
        }, status=400)
    
    try:
        data = json.loads(request.body)
        rejection_reason = data.get('rejection_reason', '').strip()
        
        if not rejection_reason:
            return JsonResponse({
                'success': False, 
                'error': 'دلیل رد درخواست الزامی است.'
            }, status=400)
        
        # Reject the request
        purchase_request.reject_request(
            rejected_by_user=request.user,
            rejection_reason=rejection_reason
        )
        
        return JsonResponse({
            'success': True, 
            'message': f'درخواست خرید شماره {request_id} با موفقیت رد شد.',
            'new_status': purchase_request.get_status_display(),
            'rejected_by': str(purchase_request.rejected_by),
            'rejected_at': purchase_request.rejected_at.strftime('%Y-%m-%d %H:%M') if purchase_request.rejected_at else None
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False, 
            'error': 'داده های ارسالی نامعتبر است.'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'خطا در پردازش درخواست: {str(e)}'
        }, status=500)
