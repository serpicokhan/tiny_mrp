from django.shortcuts import render
from mrp.models import *
import jdatetime
from django.db import IntegrityError
from django.views.decorators.csrf import csrf_exempt
import json
from django.http import JsonResponse
from mrp.business.DateJob import *
from datetime import datetime, timedelta
from django.contrib.contenttypes.models import ContentType
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.context_processors import PermWrapper
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, ExpressionWrapper, fields
from datetime import timedelta,datetime as dt
from mrp.business.ReportUtility import *
from django.template.loader import render_to_string
from django.shortcuts import get_object_or_404
from django.views.decorators import csrf
from mrp.forms import ReportForm
from django.utils import timezone
from django.db.models import Sum, F, ExpressionWrapper, FloatField
from django.db.models import Func, F, Value, FloatField, Sum, Case, When
from django.db.models.functions import Coalesce
from datetime import datetime
from django.core.paginator import Paginator
from django.db.models import Q

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.views import View


@login_required
def daily_tolid_with_chart(request):
    
    return render(request,'mrp/report/daily_tolid.html',{})




# Custom function for MySQL JSON array length
class JsonArrayLength(Func):
    function = 'JSON_LENGTH'
    template = '%(function)s(%(expressions)s)'
    output_field = models.IntegerField()

@login_required
def daily_tolid_main(request):
    # Fetch locations, categories, and shifts
    locations = Asset.objects.filter(assetIsLocatedAt__isnull=True)
    categories = AssetCategory.objects.all().order_by('priority')
    shifts = Shift.objects.all()
    profiles=FinancialProfile.objects.all()

    # Initialize query with select_related for machine only
    productions = DailyProduction.objects.filter(production_value__gt=0).select_related(
        'machine'
    ).only(
        'dayOfIssue', 'production_value', 'wastage_value', 'machine__assetName', 'operators_data'
    ).order_by('-dayOfIssue')

    # Handle filters from GET request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    operator_id = request.GET.get('operator_data')
    machine_id = request.GET.get('machine_id')
    category_id = request.GET.get('category_id')
    shift_id = request.GET.get('shift_id',False)
    st_date,e_date=False,False
    collective = request.GET.get("collective", False)
    profile_id=request.GET.get("profile_id",False)

    operator_datas=None


    # Convert date strings using DateJob
    if start_date:
        st_date=start_date
        start_date = DateJob.getTaskDate(start_date)
        productions = productions.filter(dayOfIssue__gte=start_date)
    if end_date:
        e_date=end_date
        end_date = DateJob.getTaskDate(end_date)
        productions = productions.filter(dayOfIssue__lte=end_date)

    if( not start_date and not end_date):
        
        context = {
            'makan': locations,
            'profiles':profiles,
            'category': categories,
            'shifts':shifts


        }
        return render(request, 'mrp/report/daily_tolid_main.html',context)


    # Apply filters
    if machine_id and machine_id != '-1':
        productions = productions.filter(machine__assetIsLocatedAt__id=machine_id)
    if category_id and category_id != '-1':
        productions = productions.filter(machine__assetCategory_id=category_id)
    if shift_id and shift_id != '-1':
        productions = productions.filter(shift_id=shift_id)

    # Apply operator filter (using JSONB query for PostgreSQL, if applicable)
    # if operator_id:
    #     productions = productions.filter(operators_data__contains=[{'id': operator_id}])
   # فیلتر اپراتورها
    if operator_id and operator_id!='[]':
    #     try:
    #         # پارس کردن رشته JSON
            operator_datas = json.loads(operator_id)  # تبدیل به لیست دیکشنری
            operator_id = int(operator_datas[0]['id'])  # your variable
            # print("id:",operator_id)
            productions = productions.filter(
                Q(operators_data__icontains=f'"id":{operator_id}') | 
                Q(operators_data__icontains=f'"id":"{operator_id}"')
            )
   
    

    # Calculate summary metrics in a single query
    aggregates = productions.aggregate(
        total_production=Sum(F('production_value'), output_field=FloatField(), default=0.0),
        total_wastage=Sum(F('wastage_value'), output_field=FloatField(), default=0.0)
    )
    total_production = float(aggregates['total_production'] or 0.0)  # Handle None
    total_wastage = float(aggregates['total_wastage'] or 0.0)  # Handle None
    wastage_rate = (total_wastage / total_production * 100) if total_production > 0 else 0.0

    # Paginate at the QuerySet level
   

    # Prepare report data for the current page
    report_data = []
    if collective:
        # Group by operator and machine when collective is active
        grouped_data = {}
        for prod in productions:
            if not prod.operators_data:
                continue
            try:
                operators = (
                    prod.operators_data
                    if isinstance(prod.operators_data, list)
                    else json.loads(prod.operators_data)
                )
            except (json.JSONDecodeError, TypeError):
                continue

            machine_name = prod.machine.assetCategory if prod.machine else 'Unknown'
            production = float(prod.production_value) if prod.production_value is not None else 0.0
            wastage = float(prod.wastage_value) if prod.wastage_value is not None else 0.0
            operator_count = prod.get_operator_count()
            production_per_operator = prod.get_randeman_production()#production / operator_count if operator_count > 0 else 0.0
            wastage_per_operator = wastage / operator_count if operator_count > 0 else 0.0
            wastage_rate_row = (
                (wastage_per_operator / production_per_operator * 100)
                if production_per_operator > 0
                else 0.0
            )

            for op in operators:
                if 'name' not in op or 'id' not in op:
                    continue
                if operator_datas and str(op.get('id')) != str(operator_datas[0]['id']):
                    continue
                operator_name = op['name']
                key = (operator_name, machine_name)
                if key not in grouped_data:
                    grouped_data[key] = {
                        'production': 0.0,
                        'wastage': 0.0,
                        'count': 0
                    }
                grouped_data[key]['production'] += production_per_operator
                grouped_data[key]['wastage'] += wastage_per_operator
                grouped_data[key]['count'] = +1

        # Convert grouped data to report_data format
        for (operator_name, machine_name), data in grouped_data.items():
            total_production_op = data['production']
            total_wastage_op = data['wastage']
            wastage_rate_op = (
                (total_wastage_op / total_production_op * 100)
                if total_production_op > 0
                else 0.0
            )
            report_data.append({
                'date': '',  # No single date for grouped data
                'operator': operator_name,
                'machine': machine_name,
                'production': round(total_production_op, 2),
                'eval_36_tolid': 0.0,  # Placeholder, adjust if needed
                'wastage': round(total_wastage_op, 2),
                'wastage_rate': round(wastage_rate_op, 2),
                'op_count': data['count'],
                'randeman_production': 0.0  # Placeholder, adjust if needed
            })
    else:
        for prod in productions:

            operator_names = ['Unknown']
            operator_count = 1
            if prod.operators_data:
                try:
                    operators = prod.operators_data if isinstance(prod.operators_data, list) else json.loads(prod.operators_data)
                    # print(operators)
                    if(operator_id and operator_id!='[]'):

                        operator_names = [op['name'] for op in operators if 'name' in op and str(op.get('id')) == str(operator_datas[0]['id'])]
                    else:
                        operator_names = [op['name'] for op in operators if 'name' in op ]
                    operator_count = len(operator_names) if operator_names else 1
                except (json.JSONDecodeError, TypeError):
                    print(jdatetime.date.fromgregorian(date=prod.dayOfIssue).strftime('%Y/%m/%d'),'!!!!!!!!!!!!!!!')
                    

            production = float(prod.production_value) if prod.production_value is not None else 0.0
            ##hamgen 36
            production_36 = 0#float(prod.eval_36_tolid_op_count()) if prod.eval_36_tolid_op_count() is not None else 0.0
            wastage = float(prod.wastage_value) if prod.wastage_value is not None else 0.0
            production_per_operator = production / prod.get_operator_count()
            # production_per_operator_36 = production_36 / operator_count if operator_count > 0 else 0.0
            wastage_per_operator = wastage / operator_count if operator_count > 0 else 0.0
            wastage_rate_row = (wastage_per_operator / production_per_operator * 100) if production_per_operator > 0 else 0.0
            # print(jdatetime.date.fromgregorian(date=prod.dayOfIssue).strftime('%Y/%m/%d'),'!!!!!!!!!!!!!!!')
            # print(operator_names)
            

            for operator_name in operator_names:

            
                # if operators_name in 
                report_data.append({
                    'date': jdatetime.date.fromgregorian(date=prod.dayOfIssue).strftime('%Y/%m/%d'),
                    'operator': operator_name,
                    'machine': prod.machine.assetName if prod.machine else 'Unknown',
                    'production': production_per_operator,
                    'eval_36_tolid':production_36,
                    'wastage': wastage_per_operator,
                    'wastage_rate': wastage_rate_row,
                    'op_count':prod.get_operator_count(),
                    'randeman_production':prod.get_randeman_production()
                })
    paginator = Paginator(report_data, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    # Chart data (optional)
    chart_data = {}
   

   
    context = {
        'makan': locations,
        'profiles':profiles,
        'category': categories,
        'shifts': shifts,
        # 'report_data': report_data,
        'total_production': round(total_production, 2),
        'total_wastage': round(total_wastage, 2),
        'wastage_rate': round(wastage_rate, 2),
        
        'page_obj': page_obj,
        'operator_data': operator_datas if operator_id and operator_id!='[]' else None,
        # 'operators_id':request.GET.get('operator_data') if request.GET.get('operator_data') and request.GET.get('operator_data')!='[]' else None,
        'start_date':start_date,
        'end_date':end_date,
        'shift_id':int(shift_id) if shift_id else False,
        'collective':'checked' if collective else '',
        'machin_id':int(machine_id),
        'shift_id':int(shift_id) if shift_id else None,
        'category_id':int(category_id),
        'profile_id':int(profile_id)

    }

    return render(request, 'mrp/report/daily_tolid_main.html', context)







def production_chart_with_table(request):
    date_str = request.GET.get('date',False) # Modify these dates as needed

    if(not date_str):
        # Assume 'date' is passed as 'YYYY-MM-DD' format from the front end
        date_str=DailyProduction.objects.order_by('-dayOfIssue').first().dayOfIssue
    else:
        date_str=DateJob.getTaskDate(date_str)

    production_data={}
    data=[]

    # date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    
    # Query to get sum of production_value for each machine for the given date
    shifts=Shift.objects.all()
    for i in shifts:
        production_data1 = DailyProduction.objects.filter(dayOfIssue=date_str,shift=i)\
                        .values('machine__assetName')\
                        .annotate(total_production=Sum('production_value'))\
                        .order_by('machine')
        production_data2 = DailyProduction.objects.filter(dayOfIssue=date_str,shift=i)\
                        .values('machine__assetCategory__name')\
                        .annotate(total_production=Sum('production_value'))\
                        
        data.append(
            {
        'asset_category':[item['machine__assetCategory__name'] for item in production_data2],
        'production_values2': [int(item['total_production']) for item in production_data2],

        'machines': [item['machine__assetName'] for item in production_data1],
        'production_values': [int(item['total_production']) for item in production_data1],
        'date':str(jdatetime.date.fromgregorian(date=date_str).strftime("%d-%m-%Y")),
        'lable':f'شیفت {i.name}'
        
             }
        )
   
    
    # data = {
    #     'machines': [item['machine__assetName'] for item in production_data],
    #     'production_values': [int(item['total_production']) for item in production_data],
    #     'date':str(jdatetime.date.fromgregorian(date=date_str)),
        
    # }
    
    return JsonResponse(data,safe=False)


def list_report(request,id=None):
    #
    books = Report.objects.all()
    Cat=Report.Category
    wos=ReportUtility.doPaging(request,books)
    return render(request, 'mrp/reports/main.html', {'reports': wos,'cat':Cat,'section':'list_report'})
##########################################################
def save_report_form(request, form, template_name,id=None):
    data = dict()
    if (request.method == 'POST'):
        if form.is_valid():
            form.save()
            data['form_is_valid'] = True
            books = Report.objects.all()
            wos=ReportUtility.doPaging(request,books)
            data['html_report_list'] = render_to_string('mrp/reports/partialReportList.html', {
                'reports': wos
            })
            data['form_is_valid'] = True
        else:
            data['form_is_valid'] = False

    context = {'form': form}


    data['html_report_form'] = render_to_string(template_name, context, request=request)
    return JsonResponse(data)
##########################################################


def report_delete(request, id):
    comp1 = get_object_or_404(Report, id=id)
    data = dict()
    if (request.method == 'POST'):
        comp1.delete()
        data['form_is_valid'] = True  # This is just to play along with the existing code
        companies =  Report.objects.all()
        #Tasks.objects.filter(reportId=id).update(report=id)
        data['html_report_list'] = render_to_string('mrp/report/partialReportList.html', {
            'report': companies
        })
    else:
        context = {'report': comp1}
        data['html_report_form'] = render_to_string('mrp/report/partialReportDelete.html',
            context,
            request=request,
        )
    return JsonResponse(data)

##########################################################

##########################################################
def report_create(request):
    if (request.method == 'POST'):
        form = ReportForm(request.POST)
        return save_report_form(request, form, 'mrp/reports/partialReportCreate.html')
    else:

        form = ReportForm()
        return save_report_form(request, form, 'mrp/reports/partialReportCreate.html')




##########################################################
def report_update(request, id):
    company= get_object_or_404(Report, id=id)

    if (request.method == 'POST'):
        form = ReportForm(request.POST, instance=company)
    else:
        form = ReportForm(instance=company)


    return save_report_form(request, form,"mrp/reports/partialReportUpdate.html",id)
##########################################################

##########################################################
def reportSearch(request,str):
    data=dict()
    str=str.replace('empty_','')
    str=str.replace('_',' ')
    books=[]
    # print(str,len(str),'&&&&&&&&&&&')
    if(not str):
        books=Report.objects.all()
        str='empty_'
    else:
        books = Report.objects.filter(Q(reportName__contains=str)|Q(reportDetails__contains=str))
    wos=ReportUtility.doPaging(request,books)
    data['html_report_list'] = render_to_string('mrp/reports/partialReportList.html', {
         'reports': wos,
         'perms': PermWrapper(request.user),
     })
    data['html_report_paginator'] = render_to_string('mrp/reports/partialReportPagination.html', {'reports': wos,'pageType':'reportSearch' ,'pageArg':str })
    return JsonResponse(data)
def FilterReportCategory(request,id):
    data=dict()

    books=[]

    if(id=='-1'):
        books = Report.objects.all()
    else:
         books = Report.objects.filter(reportCategory=id)
    wos=ReportUtility.doPaging(request,books)
    data['html_report_list'] = render_to_string('mrp/reports/partialReportList.html', {
         'reports': wos,'perms': PermWrapper(request.user)
     })
    # print(wos)
    data['html_report_paginator'] = render_to_string('mrp/reports/partialReportPagination.html', {'reports': wos,'pageType':'FilterReportCategory','pageArg':id})
    return JsonResponse(data)
def make_favorits_report(request,id):
    rep=Report.objects.get(id=id)
    rep.reportFav=not rep.reportFav
    rep.save()
    data=dict()
    data["form_is_valid"]=True
    return JsonResponse(data)
def show_fav_reports(request,id):
    data=dict()

    books=[]

    if(id=='1'):
        books = Report.objects.filter(reportFav=True)
    else:
         books = Report.objects.all()
    wos=ReportUtility.doPaging(request,books)
    data['html_report_list'] = render_to_string('mrp/reports/partialReportList.html', {
         'reports': wos
         ,'perms': PermWrapper(request.user)
     })
    # print(wos)
    data['html_report_paginator'] = render_to_string('mrp/reports/rep_pagination2.html', {'reports': wos})
    return JsonResponse(data)
class ProductionReportExcelExport(View):
    def get(self, request, *args, **kwargs):
        # Get the same filters from the original view
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        operator_data = request.GET.get('operator_data', '[]')
        profile_id = int(request.GET.get('profile_id', -1))
        machine_id = int(request.GET.get('machine_id', -1))
        category_id = int(request.GET.get('category_id', -1))
        shift_id = int(request.GET.get('shift_id', -1))
        collective = request.GET.get('collective', '') == 'on'

        # Apply the same filtering logic as your main view
        queryset = self.get_filtered_queryset(
            start_date, end_date, operator_data, profile_id, 
            machine_id, category_id, shift_id, collective
        )

        # Create Excel workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "گزارش تولید و ضایعات"

        # Set RTL direction for Persian text
        worksheet.sheet_view.rightToLeft = True

        # Create styles
        header_style = self.create_header_style()
        data_style = self.create_data_style()
        summary_style = self.create_summary_style()

        # Add title and summary information
        self.add_title_and_filters(worksheet, start_date, end_date, header_style)
        
        # Add summary statistics
        summary_row = self.add_summary_statistics(worksheet, queryset, summary_style)
        
        # Add table headers
        headers_row = summary_row + 3
        self.add_table_headers(worksheet, headers_row, header_style)
        
        # Add data rows
        self.add_data_rows(worksheet, queryset, headers_row + 1, data_style)
        
        # Auto-adjust column widths
        self.adjust_column_widths(worksheet)
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Generate filename with timestamp
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'production_report_{timestamp}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        workbook.save(response)
        return response

    def get_filtered_queryset(self, start_date, end_date, operator_data, profile_id, machine_id, category_id, shift_id, collective):
       
        productions = DailyProduction.objects.filter(production_value__gt=0).select_related(
            'machine'
        ).only(
            'dayOfIssue', 'production_value', 'wastage_value', 'machine__assetName', 'operators_data'
        ).order_by('-dayOfIssue')

        # Handle filters from GET request
        operator_id = operator_data
        st_date,e_date=False,False

        operator_datas=None


        # Convert date strings using DateJob
        if start_date:
            st_date=start_date
            start_date = DateJob.getTaskDate(start_date)
            productions = productions.filter(dayOfIssue__gte=start_date)
        if end_date:
            e_date=end_date
            end_date = DateJob.getTaskDate(end_date)
            productions = productions.filter(dayOfIssue__lte=end_date)

       


        # Apply filters
        if machine_id and machine_id != '-1':
            productions = productions.filter(machine__assetIsLocatedAt__id=machine_id)
        if category_id and category_id != '-1':
            productions = productions.filter(machine__assetCategory_id=category_id)
        if shift_id and shift_id != '-1':
            productions = productions.filter(shift_id=shift_id)

        # Apply operator filter (using JSONB query for PostgreSQL, if applicable)
        # if operator_id:
        #     productions = productions.filter(operators_data__contains=[{'id': operator_id}])
    # فیلتر اپراتورها
        if operator_id and operator_id!='[]':
        #     try:
        #         # پارس کردن رشته JSON
                operator_datas = json.loads(operator_id)  # تبدیل به لیست دیکشنری
                operator_id = int(operator_datas[0]['id'])  # your variable
                # print("id:",operator_id)
                productions = productions.filter(
                    Q(operators_data__icontains=f'"id":{operator_id}') | 
                    Q(operators_data__icontains=f'"id":"{operator_id}"')
                )
    
        

        # Calculate summary metrics in a single query
        aggregates = productions.aggregate(
            total_production=Sum(F('production_value'), output_field=FloatField(), default=0.0),
            total_wastage=Sum(F('wastage_value'), output_field=FloatField(), default=0.0)
        )
        total_production = float(aggregates['total_production'] or 0.0)  # Handle None
        total_wastage = float(aggregates['total_wastage'] or 0.0)  # Handle None
        wastage_rate = (total_wastage / total_production * 100) if total_production > 0 else 0.0

        # Paginate at the QuerySet level
    

        # Prepare report data for the current page
        report_data = []
        if collective:
            # Group by operator and machine when collective is active
            grouped_data = {}
            for prod in productions:
                if not prod.operators_data:
                    continue
                try:
                    operators = (
                        prod.operators_data
                        if isinstance(prod.operators_data, list)
                        else json.loads(prod.operators_data)
                    )
                except (json.JSONDecodeError, TypeError):
                    continue

                machine_name = prod.machine.assetCategory if prod.machine else 'Unknown'
                production = float(prod.production_value) if prod.production_value is not None else 0.0
                wastage = float(prod.wastage_value) if prod.wastage_value is not None else 0.0
                operator_count = prod.get_operator_count()
                production_per_operator = prod.get_randeman_production()#production / operator_count if operator_count > 0 else 0.0
                wastage_per_operator = wastage / operator_count if operator_count > 0 else 0.0
                wastage_rate_row = (
                    (wastage_per_operator / production_per_operator * 100)
                    if production_per_operator > 0
                    else 0.0
                )

                for op in operators:
                    if 'name' not in op or 'id' not in op:
                        continue
                    if operator_datas and str(op.get('id')) != str(operator_datas[0]['id']):
                        continue
                    operator_name = op['name']
                    key = (operator_name, machine_name)
                    if key not in grouped_data:
                        grouped_data[key] = {
                            'production': 0.0,
                            'wastage': 0.0,
                            'count': 0
                        }
                    grouped_data[key]['production'] += production_per_operator
                    grouped_data[key]['wastage'] += wastage_per_operator
                    grouped_data[key]['count'] = +1

            # Convert grouped data to report_data format
            for (operator_name, machine_name), data in grouped_data.items():
                total_production_op = data['production']
                total_wastage_op = data['wastage']
                wastage_rate_op = (
                    (total_wastage_op / total_production_op * 100)
                    if total_production_op > 0
                    else 0.0
                )
                report_data.append({
                    'date': '',  # No single date for grouped data
                    'operator': operator_name,
                    'machine': machine_name,
                    'production': round(total_production_op, 2),
                    'eval_36_tolid': 0.0,  # Placeholder, adjust if needed
                    'wastage': round(total_wastage_op, 2),
                    'wastage_rate': round(wastage_rate_op, 2),
                    'op_count': data['count'],
                    'randeman_production': 0.0  # Placeholder, adjust if needed
                })
        else:
            for prod in productions:

                operator_names = ['Unknown']
                operator_count = 1
                if prod.operators_data:
                    try:
                        operators = prod.operators_data if isinstance(prod.operators_data, list) else json.loads(prod.operators_data)
                        # print(operators)
                        if(operator_id and operator_id!='[]'):

                            operator_names = [op['name'] for op in operators if 'name' in op and str(op.get('id')) == str(operator_datas[0]['id'])]
                        else:
                            operator_names = [op['name'] for op in operators if 'name' in op ]
                        operator_count = len(operator_names) if operator_names else 1
                    except (json.JSONDecodeError, TypeError):
                        print(jdatetime.date.fromgregorian(date=prod.dayOfIssue).strftime('%Y/%m/%d'),'!!!!!!!!!!!!!!!')
                        

                production = float(prod.production_value) if prod.production_value is not None else 0.0
                ##hamgen 36
                production_36 = 0#float(prod.eval_36_tolid_op_count()) if prod.eval_36_tolid_op_count() is not None else 0.0
                wastage = float(prod.wastage_value) if prod.wastage_value is not None else 0.0
                production_per_operator = production / prod.get_operator_count()
                # production_per_operator_36 = production_36 / operator_count if operator_count > 0 else 0.0
                wastage_per_operator = wastage / operator_count if operator_count > 0 else 0.0
                wastage_rate_row = (wastage_per_operator / production_per_operator * 100) if production_per_operator > 0 else 0.0
                # print(jdatetime.date.fromgregorian(date=prod.dayOfIssue).strftime('%Y/%m/%d'),'!!!!!!!!!!!!!!!')
                # print(operator_names)
                

                for operator_name in operator_names:

                
                    # if operators_name in 
                    report_data.append({
                        'date': jdatetime.date.fromgregorian(date=prod.dayOfIssue).strftime('%Y/%m/%d'),
                        'operator': operator_name,
                        'machine': prod.machine.assetName if prod.machine else 'Unknown',
                        'production': production_per_operator,
                        'eval_36_tolid':production_36,
                        'wastage': wastage_per_operator,
                        'wastage_rate': wastage_rate_row,
                        'op_count':prod.get_operator_count(),
                        'randeman_production':prod.get_randeman_production()
                    })
        return report_data
    def create_header_style(self):
        return {
            'font': Font(name='B Nazanin', size=12, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def create_data_style(self):
        return {
            'font': Font(name='Tahoma', size=10),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def create_summary_style(self):
        return {
            'font': Font(name='Tahoma', size=11, bold=True),
            'fill': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def add_title_and_filters(self, worksheet, start_date, end_date, header_style):
        # Add title
        worksheet['A1'] = 'گزارش تولید و ضایعات'
        worksheet.merge_cells('A1:I1')
        cell = worksheet['A1']
        cell.font = Font(name='Tahoma', size=16, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add filter information
        filter_row = 3
        if start_date or end_date:
            date_range = f"بازه زمانی: {start_date} تا {end_date}"
            worksheet[f'A{filter_row}'] = date_range
            worksheet[f'A{filter_row}'].font = Font(name='Tahoma', size=10, italic=True)
            filter_row += 1
        
        # Add export timestamp
        worksheet[f'A{filter_row}'] = f"تاریخ تهیه گزارش: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        worksheet[f'A{filter_row}'].font = Font(name='Tahoma', size=10, italic=True)

    def add_summary_statistics(self, worksheet, queryset, summary_style):
        # Calculate summary statistics
        total_production = sum(item['production'] or 0 for item in queryset)
        total_wastage = sum(item['wastage'] or 0 for item in queryset)
        wastage_rate = (total_wastage / total_production * 100) if total_production > 0 else 0
        
        summary_row = 6
        
        # Summary headers
        worksheet[f'A{summary_row}'] = 'خلاصه آمار'
        worksheet.merge_cells(f'A{summary_row}:D{summary_row}')
        cell = worksheet[f'A{summary_row}']
        for style_key, style_value in summary_style.items():
            setattr(cell, style_key, style_value)
        
        # Summary data
        summary_row += 1
        summary_data = [
            ('تعداد کل تولید', total_production, 'عدد'),
            ('ضایعات کل', total_wastage, 'عدد'),
            ('نرخ ضایعات', f'{wastage_rate:.2f}%', 'از کل تولید')
        ]
        
        for i, (label, value, unit) in enumerate(summary_data, 1):
            col_start = (i - 1) * 3 + 1
            worksheet.cell(summary_row, col_start, label)
            worksheet.cell(summary_row, col_start + 1, value)
            worksheet.cell(summary_row, col_start + 2, unit)
            
            # Apply styles
            for col in range(col_start, col_start + 3):
                cell = worksheet.cell(summary_row, col)
                for style_key, style_value in summary_style.items():
                    setattr(cell, style_key, style_value)
        
        return summary_row + 1

    def add_table_headers(self, worksheet, row, header_style):
        headers = [
            'تاریخ', 'اپراتور', 'تعداد op', 'دستگاه', 'تولید', 
            'همگن 36', 'تعداد ضایعات', 'نرخ ضایعات', 'عملیات'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row, col, header)
            for style_key, style_value in header_style.items():
                setattr(cell, style_key, style_value)

    def add_data_rows(self, worksheet, queryset, start_row, data_style):
        for row_idx, item in enumerate(queryset, start_row):
            data = [
                item.date if hasattr(item,'date') else '',
                str(item['operator']) if hasattr(item,'operator') else '',
                item['op_count'] or 0,
                str(item['machine']) if hasattr(item,'machine') else '',
                item['production'] or 0,
                round(item['eval_36_tolid'] or 0),
                item['wastage'] or 0,
                f"{round(item['wastage_rate'] or 0)}%" if hasattr(item, 'wastage_rate') else '0%',
                item['randeman_production'] or ''
            ]
            
            for col, value in enumerate(data, 1):
                cell = worksheet.cell(row_idx, col, value)
                for style_key, style_value in data_style.items():
                    setattr(cell, style_key, style_value)
                
                # Apply special formatting for waste rate
                if col == 8:  # Waste rate column
                    try:
                        rate = float(str(value).replace('%', ''))
                        if rate < 1:
                            cell.font = Font(name='Tahoma', size=10, color='28A745')  # Green
                        elif rate < 5:
                            cell.font = Font(name='Tahoma', size=10, color='FFC107')  # Yellow
                        else:
                            cell.font = Font(name='Tahoma', size=10, color='DC3545')  # Red
                    except (ValueError, AttributeError):
                        pass

    def adjust_column_widths(self, worksheet):
        # Auto-adjust column widths
        column_widths = [15, 20, 12, 20, 15, 15, 18, 15, 20]
        
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width