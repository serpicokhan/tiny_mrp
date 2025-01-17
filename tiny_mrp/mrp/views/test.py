from django.shortcuts import render
from mrp.models import *
import jdatetime
from django.db import IntegrityError
from django.views.decorators.csrf import csrf_exempt
import json
from django.http import JsonResponse
from mrp.business.DateJob import *
from datetime import datetime, timedelta
from django.contrib.contenttypes.models import ContentType
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.context_processors import PermWrapper
from django.contrib.auth.decorators import login_required
from mrp.business.tolid_util import *
from django.template.loader import render_to_string
from mrp.forms import HeatsetMetrajForm
from django.db import IntegrityError
import subprocess
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Max
from django.contrib.auth.decorators import permission_required
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.db.models import Q
from mrp.utils import utilMonth
import csv
from openpyxl import Workbook

def backup_database(request):
    # Define your database credentials and output file's path
   # Define your database credentials and output file's path
    db_name = 'mrp758_lois'
    db_user = 'mrp758_lois'  # Default XAMPP MySQL user
    output_file = 'file102.sql'  # Ensure you use double backslashes on Windows or raw string

    # Full path to the mysqldump executable in the XAMPP installation
    mysqldump_path = 'C:\\xampp\\mysql\\bin\\mysqldump.exe'

    # Command to backup MySQL database without a password
    command = f'"{mysqldump_path}" -u {db_user} {db_name} > {output_file}'

    try:
        # Execute the command
        subprocess.run(command, shell=True, check=True)

        # return HttpResponse("Database backup was successful.")
        with open(output_file, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/sql")
            response['Content-Disposistion'] = 'attachment; filename=' + os.path.basename(output_file) +'.sql'
            return response
    except subprocess.CalledProcessError:
        return HttpResponse("Failed to backup database.")
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}")
@login_required
def get_daily_amar(request):
    asset_category = AssetCategory.objects.all().order_by('priority')
    dayOfIssue=request.GET.get('event_id',datetime.datetime.now())
    # print(dayOfIssue,'!!!!!!!!!!!!!!!!!!')
    date_object = datetime.datetime.strptime(dayOfIssue, '%Y-%m-%d')

    next_day = date_object + timedelta(days=1)

# Calculate previous day
    previous_day = date_object - timedelta(days=1)
    machines=Asset.objects.filter(assetTypes=3)
    shift_id=request.GET.get("shift_id",False)
    print(shift_id)
    if(not shift_id):
        shift_id=1
    print("shiftid",shift_id)
    
    # heatsets=Asset.objects.filter(assetCategory__id=8)
    shift=Shift.objects.all()
    machines_with_formulas = []
    machines_with_formulas2 = []
    # for s in shift:
    s=shift_id
    for machine in machines:
        try:
            formula = Formula.objects.get(machine=machine)
            speedformula = SpeedFormula.objects.get(machine=machine)
            amar=DailyProduction.objects.get(machine=machine,dayOfIssue=dayOfIssue,shift=s)
            machines_with_formulas.append({'machine': machine,'vahed':machine.assetVahed, 'formula': formula.formula,'speedformula':speedformula.formula,'amar':amar,'shift':s,'shift_id':s})
            # else:
            #     machines_with_formulas.append({'machine': machine, 'formula': formula.formula,'speed':0,'nomre':0,'speedformula':speedformula.formula})


        except Formula.DoesNotExist:
            machines_with_formulas.append({'machine': machine, 'formula': None,'formula': 0,'speed':0,'nomre':0,'shift':s,'shift_id':s})
        except SpeedFormula.DoesNotExist:
            machines_with_formulas.append({'machine': machine, 'formula': None,'formula': 0,'speed':0,'nomre':0,'speedformula':0,'shift':s,'shift_id':s})
        except DailyProduction.DoesNotExist:
            machines_with_formulas.append({'machine': machine,'vahed':machine.assetVahed, 'formula': formula.formula,'speed':0,'nomre':0,'speedformula':speedformula.formula,'shift':s,'shift_id':s})

    # for s in shift:
    #     for machine in heatsets:
    #         try:
    #             formula = Formula.objects.get(machine=machine)
    #             speedformula = SpeedFormula.objects.first()
    #             amar=DailyProduction.objects.get(machine=machine,dayOfIssue=dayOfIssue,shift=s)
    #             machines_with_formulas2.append({'machine': machine, 'formula': formula.formula,'speedformula':speedformula.formula,'amar':amar,'shift':s})
    #         except Exception as ex:
    #             print(ex)

    # print("here")
    return render(request,"mrp/tolid/daily_details_aria.html",{'heatsets':machines_with_formulas2,'machines':machines_with_formulas,'cat_list':asset_category,'shifts':shift,'next_date':next_day.strftime('%Y-%m-%d'),'prev_date':previous_day.strftime('%Y-%m-%d'),'today':jdatetime.date.fromgregorian(date=date_object),'title':'آمار روزانه','shift_id':int(s)})

@login_required
def index(request):
    has_permission = request.user.has_perm('mrp.can_view_dashboard')
    if(has_permission):
       return HttpResponseRedirect(reverse('list_dashboard'))
    elif(request.user.has_perm('mrp.can_operator_purchase')):
       return HttpResponseRedirect(reverse('list_purchase_req_detail'))

    else:
       return HttpResponseRedirect(reverse('register_daily_amar'))


@login_required
def register_daily_amar(request):
    machines=Asset.objects.filter(assetTypes=3).order_by("id","assetVahed")
    date_object=datetime.datetime.now()
    next_day = date_object + timedelta(days=1)
    asset_category = AssetCategory.objects.all().order_by('priority')
    # moshakhasat=EntryForm.objects.all()
    entry_forms = EntryForm.objects.prefetch_related('asset_details__asset_category')

    # Create a list of dictionaries in the desired format
    moshakhasat = []
    # entry_data=None
    # asset_details=None
    for entry in entry_forms:
        entry_data = entry
        
        
        asset_details = [
            {
                "id": asset.id,
                "asset_category": asset.asset_category.name,  # Assuming AssetCategory2 has a `name` field
                "asset_category_id": asset.asset_category.id,  
                "nomre": asset.nomre,
                "speed": asset.speed,
            }
            for asset in entry.asset_details.all()
        ]
        try:
            moshakhasat.append({
                    "entry": entry_data,
                    "assetdetail": json.dumps(asset_details),
                })
        except:
            pass
        # print("!",len(moshakhasat))
        # except Exception as ex:
        #     print(ex)
        # print(moshakhasat)
    # Calculate previous day
    previous_day = date_object - timedelta(days=1)
    shift_id=request.GET.get('shift_id',False)
    selected_date=request.GET.get('selected_date',False)
    if(not shift_id):
        shift_id=Shift.objects.first().id
    shift=Shift.objects.all()

    machines_with_formulas = []
    for machine in machines:
        try:
            speed=DailyProduction.objects.filter(machine=machine).last()
            nomre=DailyProduction.objects.filter(machine=machine).last()
            vahed=DailyProduction.objects.filter(machine=machine).last()

            
            formula = Formula.objects.get(machine=machine)
            speedformula = SpeedFormula.objects.get(machine=machine)
            mydict={}
            mydict["machin"]=machine
            mydict["formula"]=formula.formula
            if(speed):
                mydict["speed"]=speed.speed
            else:
                mydict["speed"]=0
            if(nomre):
                mydict["nomre"]=nomre.nomre
            else:
                mydict["nomre"]=0
            if(vahed):
                mydict["vahed"]=nomre.vahed
            else:
                mydict["vahed"]=machine.assetVahed

            if(speed):
                machines_with_formulas.append({'machine': machine, 'formula': formula.formula,'speed':speed.speed,'nomre':speed.nomre,'vahed':machine.assetVahed,'speedformula':speedformula.formula,'max':"{:.0f}".format(speed.eval_max_tolid())})
            else:
                machines_with_formulas.append({'machine': machine, 'formula': formula.formula,'speed':1,'vahed':machine.assetVahed,'nomre':0,'speedformula':speedformula.formula})


        except Formula.DoesNotExist:
            machines_with_formulas.append({'machine': machine,'formula': 0,'speed':0,'nomre':0,'vahed':machine.assetVahed})
        except SpeedFormula.DoesNotExist:
            machines_with_formulas.append({'machine': machine,'formula': formula.formula,'speed':0,'nomre':0,'speedformula':0,'vahed':machine.assetVahed})
        except DailyProduction.DoesNotExist:
            machines_with_formulas.append({'machine': machine, 'formula': formula.formula,'speed':0,'nomre':0,'speedformula':speedformula.formula,'vahed':machine.assetVahed})

    return render(request,"mrp/tolid/details_aria.html",{'selected_date':selected_date,'machines':machines_with_formulas,'cat_list':asset_category,'shifts':shift,'title':'ورود داده های روزانه','prev_date':previous_day.strftime('%Y-%m-%d'),'next_date':next_day.strftime('%Y-%m-%d'),'shift_id':int(shift_id),'moshakhastat':moshakhasat})
@login_required
def tolid_heatset(request):
    machines=Asset.objects.filter(assetCategory__id=8)
    date_object=datetime.datetime.now()
    next_day = date_object + timedelta(days=1)


# Calculate previous day
    previous_day = date_object - timedelta(days=1)
    shift=Shift.objects.all()
    machines_with_formulas = []
    for machine in machines:
        try:
            speed=DailyProduction.objects.filter(machine=machine).last()
            nomre=DailyProduction.objects.filter(machine=machine).last()
            formula = Formula.objects.get(machine=machine)
            speedformula = SpeedFormula.objects.get(machine=machine)
            mydict={}
            mydict["machin"]=machine
            mydict["formula"]=formula.formula
            if(speed):
                mydict["speed"]=speed.speed
            else:
                mydict["speed"]=0
            if(nomre):
                mydict["nomre"]=nomre.nomre
            else:
                mydict["nomre"]=0

            if(speed):
                machines_with_formulas.append({'machine': machine, 'formula': formula.formula,'speed':speed.speed,'nomre':speed.nomre,'speedformula':speedformula.formula,'max':"{:.0f}".format(speed.eval_max_tolid())})
            else:
                machines_with_formulas.append({'machine': machine, 'formula': formula.formula,'speed':0,'nomre':0,'speedformula':speedformula.formula})


        except Formula.DoesNotExist:
            machines_with_formulas.append({'machine': machine, 'formula': None,'formula': 0,'speed':0,'nomre':0})
        except SpeedFormula.DoesNotExist:
            machines_with_formulas.append({'machine': machine, 'formula': None,'formula': 0,'speed':0,'nomre':0,'speedformula':0})
        except DailyProduction.DoesNotExist:
            machines_with_formulas.append({'machine': machine, 'formula': formula.formula,'speed':0,'nomre':0,'speedformula':speedformula.formula})

    return render(request,"mrp/tolid/heatset_details.html",{'machines':machines_with_formulas,'shifts':shift,'title':'ورود داده های روزانه','prev_date':previous_day.strftime('%Y-%m-%d'),'next_date':next_day.strftime('%Y-%m-%d')})

@csrf_exempt
def saveAmarTableInfo(request):
    # print(request.body)
    # print(request.POST)
    data2 = json.loads(request.body)
    data=dict()
    mydate=""#DateJob.getTaskDate(i["dayOfIssue"].replace('/','-'))
    m_zaye=None
    # persian_date=""
    # print("********")
    for table_name, table_data in data2.items():
        for i in table_data:
            m=Asset.objects.get(id=int(i["machine"]))
            s=Shift.objects.get(id=int(i["shift"]))
            d=None
            mydate=DateJob.getTaskDate(i["dayOfIssue"].replace('/','-'))


            if(i["id"]!="0"):

                d=DailyProduction.objects.filter(id=i["id"])
            else:
                d=DailyProduction.objects.filter(machine=m,shift=s,dayOfIssue=DateJob.getTaskDate(i["dayOfIssue"].replace('/','-')))

            if(d.count()>0):
                x=d[0]
                x.machine=m
                x.shift=s
                x.dayOfIssue=DateJob.getTaskDate(i["dayOfIssue"].replace('/','-'))
                # if(s.id==1):
                #     print(i)
                #     print('!!!!!!!!',i["speed"],i["id"],s.id)
                x.speed=i["speed"]
                x.nomre=i["nomre"]
                x.counter1=float(i["counter1"])
                x.counter2=float(i["counter2"])
                x.vahed=int(i["vahed"])
                x.production_value=float(i["production_value"])
                if(i["moshakhase"]!="-1"):
                    x.moshakhase=EntryForm.objects.get(id=i["moshakhase"]) #if i["production_value"]!="-1" else None
                    m_zaye=x.moshakhase
                x.zayeat=float(i["waste"])

                try:
                    x.save()
                except IntegrityError:
                    print("برای این تاریخ مقدار از قبل وجود دارد!")
                    data["error"]="برای این تاریخ مقدار از قبل وجود دارد!"

            # print(i)
            # print(i)
            # print("********")
            else:
                amar=DailyProduction()
                # amar.shift=i["shift"]
                amar.machine=m
                amar.shift=s
                amar.dayOfIssue=DateJob.getTaskDate(i["dayOfIssue"].replace('/','-'))
                amar.speed=i["speed"]
                amar.nomre=i["nomre"]
                amar.counter1=float(i["counter1"])
                amar.counter2=float(i["counter2"])
                amar.vahed=float(i["vahed"])
                
                amar.production_value=float(i["production_value"])
                amar.zayeat=float(i["waste"])
                if(i["moshakhase"]!="-1"):
                    amar.moshakhase=EntryForm.objects.get(id=i["moshakhase"]) 
                    m_zaye=amar.moshakhase
                try:
                    amar.save()
                    print("done!!!")
                except IntegrityError as ex:
                    print(ex)
                    print("A MyModel instance with this field1 and field2 combination already exists.")
                    data["error"]="برای این تاریخ مقدار از قبل وجود دارد!"


            # print("done",amar.id)
    data=dict()
    zayeat_id=m_zaye.id if m_zaye else None
    create_zayeat_on_date(mydate,zayeat_id)
    return JsonResponse(data)

@csrf_exempt
def saveAmarHTableInfo(request):
    print("######################")
    # print(request.body)
    # print(request.POST)
    data2 = json.loads(request.body)
    data=dict()
    # print("********")
    for table_name, table_data in data2.items():
        for i in table_data:


            m=Asset.objects.get(id=int(i["machine"]))
            s=Shift.objects.get(id=int(i["shift"]))
            d=None


            if(i["id"]!="0"):
                print(i["id"])
                d=DailyProduction.objects.filter(id=i["id"])
            else:

                d=DailyProduction.objects.filter(machine=m,shift=s,dayOfIssue=DateJob.getTaskDate(i["dayOfIssue"].replace('/','-')))


            if(d.count()>0):

                x=d[0]
                x.machine=m
                x.shift=s
                x.dayOfIssue=DateJob.getTaskDate(i["dayOfIssue"].replace('/','-'))
                x.speed=int(i["speed"])
                x.nomre=i["nomre"]
                x.counter=float(i["counter"])
                x.production_value=float(i["production_value"])
                x.daf_num=float(i["daf_num"])
                x.dook_weight=float(i["dook_weight"])
                x.weight1=float(i["weight1"])
                x.weight2=float(i["weight2"])
                x.weight3=float(i["weight3"])
                x.weight4=float(i["weight4"])
                x.weight5=float(i["weight5"])
                x.net_weight=float(i["vazne_baghi"])
                z=i["data_metraj"]
                if(z):
                   if('dict' in str(type(z))):
                        # z=json.loads(i["data_metraj"])
                        print(z,z['metrajdaf1'])
                        x.metrajdaf1=z["metrajdaf1"]
                        print(x.metrajdaf1)
                        x.metrajdaf2=int(i["data_metraj"]["metrajdaf2"])
                        x.metrajdaf3=int(i["data_metraj"]["metrajdaf3"])
                        x.metrajdaf4=int(i["data_metraj"]["metrajdaf4"])
                        x.metrajdaf5=int(i["data_metraj"]["metrajdaf5"])
                        x.metrajdaf6=int(i["data_metraj"]["metrajdaf6"])
                        x.metrajdaf7=int(i["data_metraj"]["metrajdaf7"])
                        x.metrajdaf8=int(i["data_metraj"]["metrajdaf8"])
                        x.makhraj_metraj_daf=int(i["data_metraj"]["makhraj_metraj_daf"])


                    # x.metrajdaf1=int(i["data_metraj"][0])
                    # x.metrajdaf2=int(i["data_metraj"]["metrajdaf2"])
                    # x.metrajdaf3=int(i["data_metraj"]["metrajdaf3"])
                    # x.metrajdaf4=int(i["data_metraj"]["metrajdaf4"])
                    # x.metrajdaf5=int(i["data_metraj"]["metrajdaf5"])
                    # x.metrajdaf6=int(i["data_metraj"]["metrajdaf6"])
                    # x.metrajdaf7=int(i["data_metraj"]["metrajdaf7"])
                    # x.metrajdaf8=int(i["data_metraj"]["metrajdaf8"])
                    # x.makhraj_metraj_daf=int(i["data_metraj"]["makhraj_metraj_daf"])
                else:
                    x.metrajdaf1=0
                    x.metrajdaf2=0
                    x.metrajdaf3=0
                    x.metrajdaf4=0
                    x.metrajdaf5=0
                    x.metrajdaf6=0
                    x.metrajdaf7=0
                    x.metrajdaf8=0
                    x.makhraj_metraj_daf=0
                try:
                    x.save()
                except IntegrityError:
                    print("برای این تاریخ مقدار از قبل وجود دارد!")
                    data["error"]="برای این تاریخ مقدار از قبل وجود دارد!"

            # print(i)
            # print(i)
            # print("********")
            else:



                amar=DailyProduction()
                z=i["data_metraj"]
                if(z):

                     if('dict' in str(type(z))):

                        amar.metrajdaf1=i["data_metraj"]["metrajdaf1"]
                        amar.metrajdaf2=i["data_metraj"]["metrajdaf2"]
                        amar.metrajdaf3=i["data_metraj"]["metrajdaf3"]
                        amar.metrajdaf4=i["data_metraj"]["metrajdaf4"]
                        amar.metrajdaf5=i["data_metraj"]["metrajdaf5"]
                        amar.metrajdaf6=i["data_metraj"]["metrajdaf6"]
                        amar.metrajdaf7=i["data_metraj"]["metrajdaf7"]
                        amar.metrajdaf8=i["data_metraj"]["metrajdaf8"]
                        amar.makhraj_metraj_daf=i["data_metraj"]["makhraj_metraj_daf"]
                else:
                    amar.metrajdaf1=0
                    amar.metrajdaf2=0
                    amar.metrajdaf3=0
                    amar.metrajdaf4=0
                    amar.metrajdaf5=0
                    amar.metrajdaf6=0
                    amar.metrajdaf7=0
                    amar.metrajdaf8=0
                    amar.makhraj_metraj_daf=0

                # amar.shift=i["shift"]
                amar.machine=m
                amar.shift=s
                amar.dayOfIssue=DateJob.getTaskDate(i["dayOfIssue"].replace('/','-'))
                amar.speed=int(i["speed"])
                amar.nomre=i["nomre"]
                amar.counter=float(i["counter"])
                amar.production_value=float(i["production_value"])
                amar.daf_num=float(i["daf_num"])
                amar.dook_weight=float(i["dook_weight"])
                amar.weight1=float(i["weight1"])
                amar.weight2=float(i["weight2"])
                amar.weight3=float(i["weight3"])
                amar.weight4=float(i["weight4"])
                amar.weight5=float(i["weight5"])
                amar.net_weight=float(i["vazne_baghi"])
                try:
                    amar.save()
                    print("done!!!")
                except IntegrityError:
                    print("A MyModel instance with this field1 and field2 combination already exists.")
                    data["error"]="برای این تاریخ مقدار از قبل وجود دارد!"

            # print("done",amar.id)

    return JsonResponse(data)


def show_daily_amar_tolid(request):
    q=request.GET.get('date',datetime.datetime.now().date())
    date_object = datetime.datetime.strptime(q, '%Y-%m-%d')

    next_day = date_object + timedelta(days=1)


# Calculate previous day
    previous_day = date_object - timedelta(days=1)
    shifts=Shift.objects.all()
    machines=Asset.objects.filter(Q(assetTypes=3)).order_by('assetCategory__priority','assetTavali')
    machines_with_amar=[]
    m_count=1

    if(q):
        sum_randeman=0
        for index,m in enumerate(machines):

            asset_types=get_asset_count(m.assetCategory)
            shift_val=[]
            sum=0

            max_speed=1
            sum_cat=0
            for i in shifts:
                try:
                    amar=DailyProduction.objects.filter(machine=m,shift=i,dayOfIssue=q)
                    if(amar.count()>0):
                    # total_production2 = amar.aggregate(Sum('production_value'))['production_value__sum'] or 0
                        shift_val.append({'value':amar[0].production_value,'shift':i})
                        sum+=amar[0].production_value
                        max_speed=amar[0].eval_max_tolid()
                    else:
                        shift_val.append({'value':0,'shift':i})

                    # print(max_speed)


                except Exception as e:
                    shift_val.append({'value':0,'shift':i})
                    print(e)
            mx_speed=0
            if(max_speed>0):
                # print(f"{sum}/{max_speed}*{shifts.count()}")
                mx_speed=(sum/(max_speed*shifts.count()))*100
            if(m.id in (1,2,11)):
                 machines_with_amar.append({'machine':m.assetName,'shift_amar':shift_val,'css':'font-weight-bold','sum':sum,'max_speed':"{:.2f} %".format(mx_speed)})

            else:
                machines_with_amar.append({'machine':m.assetName,'shift_amar':shift_val,'sum':sum,'max_speed':"{:.2f} %".format(mx_speed)})

            if(index<len(machines)):
                sum_randeman+=mx_speed

            # print(get_sum_machine_by_date_shift(m.assetCategory,q,i))

            try:
                if(machines[index].assetCategory !=machines[index+1].assetCategory and asset_types>1):

                    x=[]
                    for i in shifts:
                        x.append({'value':get_sum_machine_by_date_shift(m.assetCategory,i,q),'shift':i})

                    machines_with_amar.append({'machine':"جمع {} ها".format(m.assetCategory) ,'css':'font-weight-bold','shift_amar':x,'sum':get_sum_machin_product_by_cat(m,q),'max_speed':"{:.2f} %".format((get_sum__speed_machine_by_category(m.assetCategory,q))*100)})
                    sum_randeman=0
            except:

                if(index==len(machines)-1 and asset_types>1):

                    x=[]
                    for i in shifts:
                        x.append({'value':get_sum_machine_by_date_shift(m.assetCategory,i,q),'shift':i})
                    machines_with_amar.append({'machine':"جمع {} ها".format(m.assetCategory) ,'css':'font-weight-bold','shift_amar':x,'sum':get_sum_machin_product_by_cat(m,q),'max_speed':"{:.2f} %".format((get_sum__speed_machine_by_category(m.assetCategory,q))*100)})
                    sum_randeman=0

                pass






    return render(request,'mrp/tolid/daily_amar_tolid.html',{'shift':shifts,'machines_with_amar':machines_with_amar,'title':'راندمان روزانه تولید','next_date':next_day.strftime('%Y-%m-%d'),'prev_date':previous_day.strftime('%Y-%m-%d'),'today':jdatetime.date.fromgregorian(date=date_object),'q':q})

def export_daily_amar_tolid(request):
    q=request.GET.get('date',datetime.datetime.now().date())
    date_object = datetime.datetime.strptime(q, '%Y-%m-%d')

    next_day = date_object + timedelta(days=1)


# Calculate previous day
    previous_day = date_object - timedelta(days=1)
    shifts=Shift.objects.all()
    machines=Asset.objects.filter(Q(assetTypes=3)).order_by('assetCategory__priority','assetTavali')
    machines_with_amar=[]
    m_count=1

    if(q):
        sum_randeman=0
        for index,m in enumerate(machines):

            asset_types=get_asset_count(m.assetCategory)
            shift_val=[]
            sum=0

            max_speed=1
            sum_cat=0
            for i in shifts:
                try:
                    amar=DailyProduction.objects.filter(machine=m,shift=i,dayOfIssue=q)
                    if(amar.count()>0):
                    # total_production2 = amar.aggregate(Sum('production_value'))['production_value__sum'] or 0
                        shift_val.append({'value':amar[0].production_value,'shift':i})
                        sum+=amar[0].production_value
                        max_speed=amar[0].eval_max_tolid()
                    else:
                        shift_val.append({'value':0,'shift':i})

                    # print(max_speed)


                except Exception as e:
                    shift_val.append({'value':0,'shift':i})
                    print(e)
            mx_speed=0
            if(max_speed>0):
                # print(f"{sum}/{max_speed}*{shifts.count()}")
                mx_speed=(sum/(max_speed*shifts.count()))*100
            if(m.id in (1,2,11)):
                 machines_with_amar.append({'machine':m.assetName,'shift_amar':shift_val,'css':'font-weight-bold','sum':sum,'max_speed':"{:.2f} %".format(mx_speed)})

            else:
                machines_with_amar.append({'machine':m.assetName,'shift_amar':shift_val,'sum':sum,'max_speed':"{:.2f} %".format(mx_speed)})

            if(index<len(machines)):
                sum_randeman+=mx_speed

            # print(get_sum_machine_by_date_shift(m.assetCategory,q,i))

            try:
                if(machines[index].assetCategory !=machines[index+1].assetCategory and asset_types>1):

                    x=[]
                    for i in shifts:
                        x.append({'value':get_sum_machine_by_date_shift(m.assetCategory,i,q),'shift':i})

                    machines_with_amar.append({'machine':"جمع {} ها".format(m.assetCategory) ,'css':'font-weight-bold','shift_amar':x,'sum':get_sum_machin_product_by_cat(m,q),'max_speed':"{:.2f} %".format((get_sum__speed_machine_by_category(m.assetCategory,q))*100)})
                    sum_randeman=0
            except:

                if(index==len(machines)-1 and asset_types>1):

                    x=[]
                    for i in shifts:
                        x.append({'value':get_sum_machine_by_date_shift(m.assetCategory,i,q),'shift':i})
                    machines_with_amar.append({'machine':"جمع {} ها".format(m.assetCategory) ,'css':'font-weight-bold','shift_amar':x,'sum':get_sum_machin_product_by_cat(m,q),'max_speed':"{:.2f} %".format((get_sum__speed_machine_by_category(m.assetCategory,q))*100)})
                    sum_randeman=0

                pass






    # return render(request,'mrp/tolid/daily_amar_tolid.html',{'shift':shifts,'machines_with_amar':machines_with_amar,'title':'راندمان روزانه تولید','next_date':next_day.strftime('%Y-%m-%d'),'prev_date':previous_day.strftime('%Y-%m-%d'),'today':jdatetime.date.fromgregorian(date=date_object)})
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Exported Data"
    shift_names = [shift.name for shift in shifts]

    # Write headers
    headers = ["نام دستگاه"] + shift_names + ["جمع"]  # Dynamic header
    sheet.append(headers)

    # Write data rows
    queryset = machines_with_amar
    for obj in queryset:
        shift_values = [k["value"] for k in obj["shift_amar"]]
        # for k in obj["shift_amar"]:
        sheet.append([obj["machine"]]+shift_values+ [obj["sum"]])

    # Create response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="exported_data.xlsx"'
    workbook.save(response)
    return response


def show_daily_amar_tolid_brief(request):
    q=request.GET.get('date',datetime.datetime.now().date())
    date_object = datetime.datetime.strptime(q, '%Y-%m-%d')

    next_day = date_object + timedelta(days=1)


# Calculate previous day
    previous_day = date_object - timedelta(days=1)
    shifts=Shift.objects.all()
    machines=Asset.objects.filter(Q(assetTypes=3)).order_by('assetCategory__priority','assetTavali')
    machines_with_amar=[]
    m_count=1

    if(q):
        sum_randeman=0
        for index,m in enumerate(machines):

            asset_types=get_asset_count(m.assetCategory)
            shift_val=[]
            sum=0

            max_speed=1
            sum_cat=0
            for i in shifts:
                try:
                    amar=DailyProduction.objects.filter(machine=m,shift=i,dayOfIssue=q)[0]
                    # total_production2 = amar.aggregate(Sum('production_value'))['production_value__sum'] or 0
                    shift_val.append({'value':amar.production_value,'shift':i})
                    sum+=amar.production_value
                    max_speed=amar.eval_max_tolid()
                    


                except Exception as e:
                    shift_val.append({'value':0,'shift':i})
            mx_speed=0
            if(max_speed>0):
                mx_speed=(sum/max_speed)*100
            if(m.id in (1,2,11)):
                 machines_with_amar.append({'machine':m.assetName,'shift_amar':shift_val,'css':'font-weight-bold','sum':sum,'max_speed':"{:.2f} %".format(mx_speed)})

            # else:
            #     machines_with_amar.append({'machine':m.assetName,'shift_amar':shift_val,'sum':sum,'max_speed':"{:.2f} %".format(mx_speed)})

            if(index<len(machines)):
                sum_randeman+=mx_speed

            # print(get_sum_machine_by_date_shift(m.assetCategory,q,i))

            try:
                if(machines[index].assetCategory !=machines[index+1].assetCategory and asset_types>1):

                    x=[]
                    for i in shifts:
                        x.append({'value':get_sum_machine_by_date_shift(m.assetCategory,i,q),'shift':i})

                    machines_with_amar.append({'machine':"{}".format(m.assetCategory) ,'css':'font-weight-bold','shift_amar':x,'sum':get_sum_machin_product_by_cat(m,q),'max_speed':"{:.2f} %".format((get_sum__speed_machine_by_category(m.assetCategory,q))*100)})
                    sum_randeman=0
            except:

                if(index==len(machines)-1 and asset_types>1):

                    x=[]
                    for i in shifts:
                        x.append({'value':get_sum_machine_by_date_shift(m.assetCategory,i,q),'shift':i})
                    machines_with_amar.append({'machine':"{}".format(m.assetCategory) ,'css':'font-weight-bold','shift_amar':x,'sum':get_sum_machin_product_by_cat(m,q),'max_speed':"{:.2f} %".format((get_sum__speed_machine_by_category(m.assetCategory,q))*100)})
                    sum_randeman=0

                pass
    return render(request,'mrp/tolid/daily_amar_tolid_brief.html',{'shift':shifts,'machines_with_amar':machines_with_amar,'title':'راندمان روزانه تولید','next_date':next_day.strftime('%Y-%m-%d'),'prev_date':previous_day.strftime('%Y-%m-%d'),'today':jdatetime.date.fromgregorian(date=date_object),'q':q})
        
###############Export to Excel #############3
def export_daily_amar_tolid_brief(request):
    q=request.GET.get('date',datetime.datetime.now().date())
    date_object = datetime.datetime.strptime(q, '%Y-%m-%d')

    next_day = date_object + timedelta(days=1)


# Calculate previous day
    previous_day = date_object - timedelta(days=1)
    shifts=Shift.objects.all()
    machines=Asset.objects.filter(Q(assetTypes=3)).order_by('assetCategory__priority','assetTavali')
    machines_with_amar=[]
    m_count=1

    if(q):
        sum_randeman=0
        for index,m in enumerate(machines):

            asset_types=get_asset_count(m.assetCategory)
            shift_val=[]
            sum=0

            max_speed=1
            sum_cat=0
            for i in shifts:
                try:
                    amar=DailyProduction.objects.filter(machine=m,shift=i,dayOfIssue=q)[0]
                    # total_production2 = amar.aggregate(Sum('production_value'))['production_value__sum'] or 0
                    shift_val.append({'value':amar.production_value,'shift':i})
                    sum+=amar.production_value
                    max_speed=amar.eval_max_tolid()
                    


                except Exception as e:
                    shift_val.append({'value':0,'shift':i})
            mx_speed=0
            if(max_speed>0):
                mx_speed=(sum/max_speed)*100
            if(m.id in (1,2,11)):
                 machines_with_amar.append({'machine':m.assetName,'shift_amar':shift_val,'css':'font-weight-bold','sum':sum,'max_speed':"{:.2f} %".format(mx_speed)})

            # else:
            #     machines_with_amar.append({'machine':m.assetName,'shift_amar':shift_val,'sum':sum,'max_speed':"{:.2f} %".format(mx_speed)})

            if(index<len(machines)):
                sum_randeman+=mx_speed

            # print(get_sum_machine_by_date_shift(m.assetCategory,q,i))

            try:
                if(machines[index].assetCategory !=machines[index+1].assetCategory and asset_types>1):

                    x=[]
                    for i in shifts:
                        x.append({'value':get_sum_machine_by_date_shift(m.assetCategory,i,q),'shift':i})

                    machines_with_amar.append({'machine':"{}".format(m.assetCategory) ,'css':'font-weight-bold','shift_amar':x,'sum':get_sum_machin_product_by_cat(m,q),'max_speed':"{:.2f} %".format((get_sum__speed_machine_by_category(m.assetCategory,q))*100)})
                    sum_randeman=0
            except:

                if(index==len(machines)-1 and asset_types>1):

                    x=[]
                    for i in shifts:
                        x.append({'value':get_sum_machine_by_date_shift(m.assetCategory,i,q),'shift':i})
                    machines_with_amar.append({'machine':"{}".format(m.assetCategory) ,'css':'font-weight-bold','shift_amar':x,'sum':get_sum_machin_product_by_cat(m,q),'max_speed':"{:.2f} %".format((get_sum__speed_machine_by_category(m.assetCategory,q))*100)})
                    sum_randeman=0

                pass





    # return render(request,'mrp/tolid/daily_amar_tolid_brief.html',{'shift':shifts,'machines_with_amar':machines_with_amar,'title':'راندمان روزانه تولید','next_date':next_day.strftime('%Y-%m-%d'),'prev_date':previous_day.strftime('%Y-%m-%d'),'today':jdatetime.date.fromgregorian(date=date_object)})
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Exported Data"

    # Write headers
    sheet.append(["نام دستگاه", "جمع"])  # Adjust fields

    # Write data rows
    queryset = machines_with_amar
    for obj in queryset:
        sheet.append([obj["machine"], obj["sum"]])

    # Create response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="exported_data.xlsx"'
    workbook.save(response)
    return response


def show_daily_analyse_tolid(request):
        q=request.GET.get('date',datetime.datetime.now().date())
        q=request.GET.get('date',datetime.datetime.now().date())
        date_object = datetime.datetime.strptime(q, '%Y-%m-%d')
        next_day = date_object + timedelta(days=1)



    # Calculate previous day
        previous_day = date_object - timedelta(days=1)
        shifts=Shift.objects.all()
        machines=Asset.objects.filter(Q(assetTypes=2)|Q(assetCategory__id=8))
        machines_with_amar=[]
        if(q):
            for index,m in enumerate(machines):
                asset_types=get_asset_count(m.assetCategory)
                shift_val=[]
                sum=0
                max_speed=0
                tolid_standard=ProductionStandard.objects.get(machine_name=m)
                for i in shifts:
                    try:


                        amar=DailyProduction.objects.filter(machine=m,shift=i,dayOfIssue=q)[0]
                        sum+=amar.production_value

                    except Exception as e:
                        print(e)
                natije=round(sum-tolid_standard.good_production_rate)
                machines_with_amar.append({'machine':m.assetName,'good':tolid_standard.good_production_rate,'mean':tolid_standard.mean_production_rate,
                'bad':tolid_standard.bad_production_rate,'real':sum,'kasre_tolid':natije})
                try:
                    if(machines[index].assetCategory !=machines[index+1].assetCategory and asset_types>1):


                        # print(sum_randeman)
                        tolid=get_sum_machin_product_by_cat(m,q)
                        good_tolid=get_good_standard_machine_by_date_category(m.assetCategory)
                        # machines_with_amar.append({'machine':"جمع {} ها".format(m.assetCategory) ,'css':'font-weight-bold','shift_amar':x,'sum':get_sum_machin_product_by_cat(m,q),'max_speed':"{:.2f} %".format((get_sum__speed_machine_by_category(m.assetCategory,q))*100)})
                        machines_with_amar.append({'machine':"جمع {} ها".format(m.assetCategory),'css':'font-weight-bold','good':good_tolid,'mean':get_mean_standard_machine_by_date_category(m.assetCategory)
                        ,'bad':get_bad_standard_machine_by_date_category(m.assetCategory),'real':tolid,'kasre_tolid':tolid-good_tolid})


                        sum_randeman=0
                except Exception as ex:
                    if(index==len(machines)-1 and asset_types>1):
                            tolid=get_sum_machin_product_by_cat(m,q)
                            good_tolid=get_good_standard_machine_by_date_category(m.assetCategory)



                            machines_with_amar.append({'machine':"جمع {} ها".format(m.assetCategory),'css':'font-weight-bold','good':good_tolid,'mean':get_mean_standard_machine_by_date_category(m.assetCategory)
                            ,'bad':get_bad_standard_machine_by_date_category(m.assetCategory),'real':tolid,'kasre_tolid':tolid-good_tolid})





        return render(request,'mrp/tolid/daily_analyse_tolid.html',{'machines_with_amar':machines_with_amar,'title':'تحلیل روزانه تولید','next_date':next_day.strftime('%Y-%m-%d'),'prev_date':previous_day.strftime('%Y-%m-%d'),'today':jdatetime.date.fromgregorian(date=date_object)})
def calendar_main(request):
    return render(request,'mrp/tolid/calendar_main.html',{})
def calendar_randeman(request):
    return render(request,'mrp/tolid/calendar_randeman.html',{'title':'راندمان روزانه'})
def calendar_randeman_brief(request):
    return render(request,'mrp/tolid/calendar_randeman_brief.html',{'title':'راندمان روزانه'})
def calendar_tahlil(request):
    return render(request,'mrp/tolid/calendar_tahlil.html',{'title':'تحلیل روزانه'})
def get_tolid_calendar_info(request):
    data=[]
    user_info=DailyProduction.objects.values_list('dayOfIssue').distinct()
    print(user_info)
    for i in user_info:
        z=get_sum_vaz_zayeat_by_date(i[0])
        data.append({'title': "آمار روزانه",\
                'start': i[0],\
                 'color': '#53c797',\
                'id':i[0]})
        data.append({'title': "جمع ضایعات : {}".format(round(z,0)),\
                'start': i[0],\
                 'color': 'red',\
                'id':i[0]})

    return JsonResponse(data,safe=False)
@csrf_exempt
def move_tolid_calendar_info(request):
    start=request.POST.get("start",datetime.datetime.now())
    end=request.POST.get("end",datetime.datetime.now())

    # datetime_obj = datetime.fromisoformat(start.replace('Z', '+00:00'))

    # Extract the date part and format it as 'YYYY-MM-DD'
    # start = datetime_obj.strftime('%Y-%m-%d')
    # datetime_obj = datetime.fromisoformat(end.replace('Z', '+00:00'))

    # Extract the date part and format it as 'YYYY-MM-DD'
    # end = datetime_obj.strftime('%Y-%m-%d')
    data=dict()
    if(request.method=="POST"):
        daily_amar=DailyProduction.objects.filter(dayOfIssue=start)
        zayeat=ZayeatVaz.objects.filter(dayOfIssue=start)
        with transaction.atomic():
            for i in daily_amar:
                i.dayOfIssue=end
                i.save()
            for i in zayeat:
                i.dayOfIssue=end
                i.save()
            data["success"]="success"
    return JsonResponse(data)

def get_randeman_calendar_info(request):
    data=[]
    user_info=DailyProduction.objects.values_list('dayOfIssue').distinct()
    
    for i in user_info:
        z=get_sum_vaz_zayeat_by_date(i[0])
        data.append({'title': "راندمان روزانه",\
                'start': i[0],\
                 'color': '#fb3',\
                'id':i[0]})
        data.append({'title': "جمع ضایعات : {}".format(round(z,0)),\
                'start': i[0],\
                 'color': 'red',\
                'id':i[0]})

    return JsonResponse(data,safe=False)
def get_tahlil_calendar_info(request):
    data=[]
    user_info=DailyProduction.objects.values_list('dayOfIssue').distinct()

    for i in user_info:
        z=get_sum_vaz_zayeat_by_date(i[0])
        data.append({'title': 'تحلیل روزانه',\
                'start': i[0],\
                 'color': '#a6c',\
                'id':i[0]})
        data.append({'title': "جمع ضایعات : {}".format(round(z,0)),\
                'start': i[0],\
                 'color': 'red',\
                'id':i[0]})
    return JsonResponse(data,safe=False)
def list_formula(request):
    formulas=Formula.objects.all()
    return render(request,"mrp/formula/formulaList.html",{'formulas':formulas,'title':'لیست فرمولهای تولید'})
def list_speed_formula(request):
    formulas=SpeedFormula.objects.all()
    return render(request,"mrp/speed_formula/formulaList.html",{'formulas':formulas,'title':'لیست فرمولهای سرعت'})

def monthly_detaild_report(request):
    days=[]
    shift=Shift.objects.all()
    asset_category = AssetCategory.objects.all().order_by('priority')

    current_date_time2 = jdatetime.datetime.now()
    current_year=current_date_time2.year
    j_month=request.GET.get('month',current_date_time2.month)

    j_year=int(request.GET.get('year',current_year))
    current_date_time = jdatetime.date(j_year, int(j_month), 1)
    current_jalali_date = current_date_time





    if current_jalali_date.month == 12:
        first_day_of_next_month = current_jalali_date.replace(day=1, month=1, year=j_year + 1)
    else:
        first_day_of_next_month = current_jalali_date.replace(day=1, month=current_jalali_date.month + 1)


    num_days = (first_day_of_next_month - jdatetime.timedelta(days=1)).day
    cat_list=[]
    for cats in asset_category:
        sh_list=[]

        days=[]
        for day in range(1,num_days+1):
            product={}
            j_date=jdatetime.date(j_year,current_jalali_date.month,day)
            # print(j_date,'!!!!!!!!!!!!!')
            for sh in shift:
                product[sh.id]=get_sum_machine_by_date_shift(cats,sh,j_date.togregorian())
                print(product[sh.id])
            days.append({'cat':cats,'date':"{0}/{1}/{2}".format(j_year,current_jalali_date.month,day),'day_of_week':DateJob.get_day_of_week(j_date),'product':product})
        product={}
        start=jdatetime.date(j_year,current_jalali_date.month,1)
        end=jdatetime.date(j_year,current_jalali_date.month,num_days)
        for sh in shift:
            product[sh.id]=get_monthly_machine_by_date_shift(cats,sh,start.togregorian(),end.togregorian())
        days.append({'cat':cats,'date':"",'day_of_week':'جمع','product':product})
        failure_days={}
        for sh in shift:
            failure_days[sh.id]=get_day_machine_failure_monthly_shift(cats,sh,start.togregorian(),end.togregorian())

        total_day_per_shift={}
        for sh in shift:
            total_day_per_shift[sh.id]=num_days-failure_days[sh.id]
        days.append({'cat':cats,'date':"",'day_of_week':'روز کاری','product':total_day_per_shift})
        mean_day_per_shift={}
        for sh in shift:
            mean_day_per_shift[sh.id]=product[sh.id]/total_day_per_shift[sh.id]

        days.append({'cat':cats,'date':"",'day_of_week':'میانگین','product':mean_day_per_shift})


        cat_list.append({'cat':cats,'shift_val':days})
        # print(cat_list)

    return render(request,'mrp/tolid/monthly_detailed.html',{'cats':asset_category,'title':'آمار ماهانه','cat_list':cat_list,'shift':shift,'month':j_month,'year':j_year})

def monthly_detaild_export(request):
    days=[]
    shift=Shift.objects.all()
    asset_category = AssetCategory.objects.all().order_by('priority')

    current_date_time2 = jdatetime.datetime.now()
    current_year=current_date_time2.year
    j_month=request.GET.get('month',current_date_time2.month)

    j_year=int(request.GET.get('year',current_year))
    current_date_time = jdatetime.date(j_year, int(j_month), 1)
    current_jalali_date = current_date_time





    if current_jalali_date.month == 12:
        first_day_of_next_month = current_jalali_date.replace(day=1, month=1, year=j_year + 1)
    else:
        first_day_of_next_month = current_jalali_date.replace(day=1, month=current_jalali_date.month + 1)


    num_days = (first_day_of_next_month - jdatetime.timedelta(days=1)).day
    cat_list=[]
    for cats in asset_category:
        sh_list=[]

        days=[]
        for day in range(1,num_days+1):
            product={}
            j_date=jdatetime.date(j_year,current_jalali_date.month,day)
            # print(j_date,'!!!!!!!!!!!!!')
            for sh in shift:
                product[sh.id]=get_sum_machine_by_date_shift(cats,sh,j_date.togregorian())
                # print(product[sh.id])
            days.append({'cat':cats,'date':"{0}/{1}/{2}".format(j_year,current_jalali_date.month,day),'day_of_week':DateJob.get_day_of_week(j_date),'product':product})
        product={}
        start=jdatetime.date(j_year,current_jalali_date.month,1)
        end=jdatetime.date(j_year,current_jalali_date.month,num_days)
        for sh in shift:
            product[sh.id]=get_monthly_machine_by_date_shift(cats,sh,start.togregorian(),end.togregorian())
        days.append({'cat':cats,'date':"",'day_of_week':'جمع','product':product})
        failure_days={}
        for sh in shift:
            failure_days[sh.id]=get_day_machine_failure_monthly_shift(cats,sh,start.togregorian(),end.togregorian())

        total_day_per_shift={}
        for sh in shift:
            total_day_per_shift[sh.id]=num_days-failure_days[sh.id]
        days.append({'cat':cats,'date':"",'day_of_week':'روز کاری','product':total_day_per_shift})
        mean_day_per_shift={}
        for sh in shift:
            mean_day_per_shift[sh.id]=product[sh.id]/total_day_per_shift[sh.id]

        days.append({'cat':cats,'date':"",'day_of_week':'میانگین','product':mean_day_per_shift})


        cat_list.append({'cat':cats,'shift_val':days})
        # print(cat_list)

    # return render(request,'mrp/tolid/monthly_detailed.html',{'cats':asset_category,'title':'آمار ماهانه','cat_list':cat_list,'shift':shift,'month':j_month,'year':j_year})
    workbook = Workbook()
    sheet = workbook.active
    # sheet.title = "Exported Data"
    shift_names = [shif.name for shif in shift]

    # # Write headers
    # headers = ["نام دستگاه"] + shift_names + ["جمع"]  # Dynamic header
    # sheet.append(headers)
    for sheet_name in cat_list:
        # Create a new sheet for each dataset
        sheet = workbook.create_sheet(title=sheet_name["cat"].name)

        # Add headers (you can customize these per dataset if needed)
        headers = ["روز هفته", "تاریخ"]+shift_names+ ["جمع"]  # Adjust fields as per your models
        sheet.append(headers)
        for day in sheet_name["shift_val"]:
            if sheet_name["cat"] == day["cat"]:
                print(day["product"])
                sheet.append([day["day_of_week"],day["date"],day["product"][1],day["product"][2],day["product"][3],(day["product"][2]+day["product"][2]+day["product"][3])])#,day["product"][shift[0].id]])#+shift_values+ [obj["sum"]])



    # Write data rows
    # queryset = machines_with_amar
    # for obj in queryset:
    #     shift_values = [k["value"] for k in obj["shift_amar"]]
    #     # for k in obj["shift_amar"]:
    #     sheet.append([obj["machine"]]+shift_values+ [obj["sum"]])

    # # Create response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="exported_data.xlsx"'
    workbook.save(response)
    return response


def monthly_brief_report(request):
    shifts=Shift.objects.all()
    asset_cats=AssetCategory.objects.all().order_by('priority')
    current_date_time2 = jdatetime.datetime.now()

    current_year=current_date_time2.year
    j_month=request.GET.get('month',current_date_time2.month)

    j_year=int(request.GET.get('year',current_year))
    current_date_time = jdatetime.date(j_year, int(j_month), 1)
    current_jalali_date = current_date_time
    if current_jalali_date.month == 12:
        first_day_of_next_month = current_jalali_date.replace(day=1, month=1, year=j_year + 1)
    else:
        first_day_of_next_month = current_jalali_date.replace(day=1, month=current_jalali_date.month + 1)


    num_days = (first_day_of_next_month - jdatetime.timedelta(days=1)).day

    totals=[]
    sum={}
    for sh in shifts:
        sum[sh.id]=0

    for cats in asset_cats:
            product={}
            start=jdatetime.date(j_year,current_jalali_date.month,1)
            end=jdatetime.date(j_year,current_jalali_date.month,num_days)
            for sh in shifts:
                product[sh.id]=get_monthly_machine_by_date_shift(cats,sh,start.togregorian(),end.togregorian())

            # days.append({'cat':cats,'date':"",'day_of_week':'جمع','product':product})
            failure_days={}
            for sh in shifts:
                failure_days[sh.id]=get_day_machine_failure_monthly_shift(cats,sh,start.togregorian(),end.togregorian())

            total_day_per_shift={}
            for sh in shifts:

                total_day_per_shift[sh.id]=num_days-failure_days[sh.id]
            # days.append({'cat':cats,'date':"",'day_of_week':'روز کاری','product':total_day_per_shift})
            mean_day_per_shift={}
            for sh in shifts:
                # if(cats.id==9 or cats.id==10):
                #     mean_day_per_shift[sh.id]=2000
                #     sum[sh.id]+=2000
                # else:


                mean_day_per_shift[sh.id]=product[sh.id]/total_day_per_shift[sh.id]
                sum[sh.id]+=mean_day_per_shift[sh.id]


            totals.append({'cat':cats,'date':"",'day_of_week':'میانگین','product':mean_day_per_shift})




    return render(request,'mrp/tolid/monthly_brief.html',{'cats':totals,'sum':sum,'shift':shifts,'title':'آمار ماهانه کلی','month':j_month,'year':j_year})


def monthly_brief_export(request):
    shifts=Shift.objects.all()
    asset_cats=AssetCategory.objects.all().order_by('priority')
    current_date_time2 = jdatetime.datetime.now()

    current_year=current_date_time2.year
    j_month=request.GET.get('month',current_date_time2.month)

    j_year=int(request.GET.get('year',current_year))
    current_date_time = jdatetime.date(j_year, int(j_month), 1)
    current_jalali_date = current_date_time
    if current_jalali_date.month == 12:
        first_day_of_next_month = current_jalali_date.replace(day=1, month=1, year=j_year + 1)
    else:
        first_day_of_next_month = current_jalali_date.replace(day=1, month=current_jalali_date.month + 1)


    num_days = (first_day_of_next_month - jdatetime.timedelta(days=1)).day

    totals=[]
    sum={}
    for sh in shifts:
        sum[sh.id]=0

    for cats in asset_cats:
            product={}
            start=jdatetime.date(j_year,current_jalali_date.month,1)
            end=jdatetime.date(j_year,current_jalali_date.month,num_days)
            for sh in shifts:
                product[sh.id]=get_monthly_machine_by_date_shift(cats,sh,start.togregorian(),end.togregorian())

            # days.append({'cat':cats,'date':"",'day_of_week':'جمع','product':product})
            failure_days={}
            for sh in shifts:
                failure_days[sh.id]=get_day_machine_failure_monthly_shift(cats,sh,start.togregorian(),end.togregorian())

            total_day_per_shift={}
            for sh in shifts:

                total_day_per_shift[sh.id]=num_days-failure_days[sh.id]
            # days.append({'cat':cats,'date':"",'day_of_week':'روز کاری','product':total_day_per_shift})
            mean_day_per_shift={}
            for sh in shifts:
                # if(cats.id==9 or cats.id==10):
                #     mean_day_per_shift[sh.id]=2000
                #     sum[sh.id]+=2000
                # else:


                mean_day_per_shift[sh.id]=product[sh.id]/total_day_per_shift[sh.id]
                sum[sh.id]+=mean_day_per_shift[sh.id]


            totals.append({'cat':cats,'date':"",'day_of_week':'میانگین','product':mean_day_per_shift})




    # return render(request,'mrp/tolid/monthly_brief.html',{'cats':totals,'sum':sum,'shift':shifts,'title':'آمار ماهانه کلی','month':j_month,'year':j_year})
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Exported Data"
    shift_names = [shift.name for shift in shifts]

    # Write headers
    headers = ["نام دستگاه"] + shift_names + ["جمع"]  # Dynamic header
    sheet.append(headers)

    # Write data rows
    queryset = totals
    for obj in queryset:
        # shift_values = [k["product"] for k in obj["product"]]
        # for k in obj["shift_amar"]:
        sheet.append([obj["cat"].name,obj["product"][1],obj["product"][2],obj["product"][3],obj["product"][1]+obj["product"][2]+obj["product"][3]])
    sum_sum=0
    for val in sum.items():
        sum_sum+=val[1]

    sheet.append(["جمع"]+[value for key, value in sum.items()]+[sum_sum])
    # Create response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="exported_data.xlsx"'
    workbook.save(response)
    return response










def list_randeman_tolid(request):
    formulas=AssetRandemanInit.objects.all()
    return render(request,"mrp/tolid_randeman/randemanList.html",{'formulas':formulas,'title':'لیست راندمان'})

def get_sum_randeman_by_shift(mah,sal,shift):
    asset_randeman_list=AssetRandemanList.objects.get(sal=sal,mah=mah)



    filtered_production = AssetRandemanPerMonth.objects.filter(
    asset_randeman_list=asset_randeman_list,  # Filter by date range

    shift=shift  # Filter by asset category n
    )
    # Calculate the sum of production_value
    sum_production_value = filtered_production.aggregate(
        total_production_value=models.Sum('tolid_value')
    )['total_production_value']

    if(not sum_production_value):
        return 0

    return sum_production_value
def get_sum_randeman(mah,sal):
    asset_randeman_list=AssetRandemanList.objects.get(sal=sal,mah=mah)



    filtered_production = AssetRandemanPerMonth.objects.filter(
    asset_randeman_list=asset_randeman_list  # Filter by date range


    )
    # Calculate the sum of production_value
    sum_production_value = filtered_production.aggregate(
        total_production_value=models.Sum('tolid_value')
    )['total_production_value']

    if(not sum_production_value):
        return 0

    return sum_production_value
def get_monthly_workbook(request):
    my_dict = {
    1: 'اول',
    2: 'دوم',
    3:'سوم'
    # Add more key-value pairs as needed
}
    mah=request.GET.get("mah",False)
    sal=request.GET.get("sal",False)
    shift_list=Shift.objects.all()
    asset_randeman_list=AssetRandemanList.objects.get(sal=sal,mah=mah)
    randeman_list=AssetRandemanPerMonth.objects.filter(asset_randeman_list=asset_randeman_list).order_by('asset_category__priority')
    profile=AssetRandemanList.objects.get(sal=sal,mah=mah).profile
    d=[]
    for i in randeman_list:
        d.append({'operator_num':AssetRandemanInit.objects.get(asset_category=i.asset_category,profile=profile).operator_count,'randeman':i})
    k=[]
    sum_randeman_tolid_kol=0
    sum_nezafat_kol=0
    sum_padash_tolid_kol=0
    sum_randeman_tolid_kol_majmu=0
    for i in shift_list:
        randeman_list=AssetRandemanList.objects.get(mah=mah,sal=sal)
        nezafat_rank=NezafatRanking.objects.get(asset_randeman_list=randeman_list,shift=i).rank
        tolid_rank=TolidRanking.objects.get(asset_randeman_list=randeman_list,shift=i).rank
        padashe_nezafat_personel=NezafatRanking.objects.get(asset_randeman_list=randeman_list,shift=i).price_personnel
        padashe_tolid_personel=TolidRanking.objects.get(asset_randeman_list=randeman_list,shift=i).price_personnel
        randeman_kol=get_sum_randeman_by_shift(mah,sal,i)
        sum=randeman_kol+padashe_nezafat_personel+padashe_tolid_personel
        k.append({'randeman_kol':randeman_kol,'shift':i,'nezafat_rank':my_dict[nezafat_rank],'tolid_rank':my_dict[tolid_rank],'padashe_nezafat':padashe_nezafat_personel,'padashe_tolid':padashe_tolid_personel,'sum':sum})
        sum_randeman_tolid_kol_majmu+=sum
        sum_padash_tolid_kol+=padashe_tolid_personel
        sum_nezafat_kol+=padashe_nezafat_personel
        sum_randeman_tolid_kol+=randeman_kol

    return render(request,'mrp/assetrandeman/finalRandemanList.html',{'shift_list':shift_list,'randeman_list':d,'randeman_kol':k,'mah':utilMonth[12-int(mah)],'sal':sal,
                            'sum_randeman_tolid_kol_majmu':sum_randeman_tolid_kol_majmu,
                            'sum_padash_tolid_kol':sum_padash_tolid_kol,'sum_nezafat_kol':sum_nezafat_kol,
                            'sum_randeman_tolid_kol':sum_randeman_tolid_kol})
def get_monthly_sarshift_workbook(request):
        my_dict = {
        1: 'اول',
        2: 'دوم',
        3:'سوم'
        # Add more key-value pairs as needed
        }
        mah=request.GET.get("mah",False)
        sal=request.GET.get("sal",False)
        shift_list=Shift.objects.all()
        k=[]

        sum_shift_randeman_tolid=0
        sum_padashe_tolid_personel=0
        sum_padashe_nezafat_personel=0
        sum_sum=0
        randeman_tolid=get_sum_randeman(mah,sal)
        for i in shift_list:
            randeman_list=AssetRandemanList.objects.get(mah=mah,sal=sal)
            nezafat_rank=NezafatRanking.objects.get(asset_randeman_list=randeman_list,shift=i).rank
            tolid_rank=TolidRanking.objects.get(asset_randeman_list=randeman_list,shift=i).rank
            padashe_nezafat_personel=NezafatRanking.objects.get(asset_randeman_list=randeman_list,shift=i).price_sarshift
            padashe_tolid_personel=TolidRanking.objects.get(asset_randeman_list=randeman_list,shift=i).price_sarshift
            padashe_tolid_23=randeman_list.profile.tolid_randeman
            randeman_kol=get_sum_randeman_by_shift(mah,sal,i)

            shift_randeman_tolid=(padashe_tolid_23*randeman_kol)/randeman_tolid
            sum_shift_randeman_tolid+=shift_randeman_tolid
            sum=padashe_nezafat_personel+padashe_tolid_personel+shift_randeman_tolid
            sum_padashe_tolid_personel+=padashe_tolid_personel
            sum_padashe_nezafat_personel+=padashe_nezafat_personel
            k.append({'randeman_kol':randeman_kol,'shift':i,'nezafat_rank':my_dict[nezafat_rank],'tolid_rank':my_dict[tolid_rank],'padashe_nezafat':padashe_nezafat_personel,'padashe_tolid':padashe_tolid_personel,'sum':sum,'shift_randeman_tolid':shift_randeman_tolid,
            })

        sum_sum=sum_shift_randeman_tolid+sum_padashe_tolid_personel+sum_padashe_nezafat_personel
        return render(request,'mrp/assetrandeman/finalSarshiftRandemanList.html',{'title':'راندمان ماهانه سر شیفت ها','k':k,
        'randeman_tolid':randeman_tolid,'shift_randeman_tolid':shift_randeman_tolid,'sum_padashe_tolid_personel':sum_padashe_tolid_personel,'sum_padashe_nezafat_personel':sum_padashe_nezafat_personel,'sum_sum':sum_sum,'sum_shift_randeman_tolid':sum_shift_randeman_tolid,'mah':utilMonth[12-int(mah)],'sal':sal})
def list_heatset_info(request):
        dayOfIssue=request.GET.get('event',False)
        if(not dayOfIssue):
            dayOfIssue=request.GET.get('event_id',datetime.datetime.now())
            date_object = DateJob.getTaskDate(dayOfIssue)
        else:
            date_object=datetime.datetime.strptime(str(dayOfIssue), '%Y-%m-%d')


        next_day = date_object + timedelta(days=1)
        data=dict()

    # Calculate previous day
        previous_day = date_object - timedelta(days=1)
        machines=Asset.objects.filter(assetCategory__id=8)
        shift=Shift.objects.all()
        machines_with_formulas = []
        for s in shift:
            for machine in machines:
                try:
                    formula = Formula.objects.get(machine=machine)
                    speedformula = SpeedFormula.objects.get(machine=machine)
                    amar=DailyProduction.objects.get(machine=machine,dayOfIssue=dayOfIssue,shift=s)
                    saved_data = {
                        'metrajdaf1': amar.metrajdaf1,
                        'metrajdaf2': amar.metrajdaf2,
                        'metrajdaf3': amar.metrajdaf3,
                        'metrajdaf4': amar.metrajdaf4,
                        'metrajdaf5': amar.metrajdaf5,
                        'metrajdaf6': amar.metrajdaf6,
                        'metrajdaf7': amar.metrajdaf7,
                        'metrajdaf8': amar.metrajdaf8,

                        'makhraj_metraj_daf': amar.makhraj_metraj_daf,
                    }
                    metraj_val = [
                        amar.metrajdaf1,
                        amar.metrajdaf2,
                        amar.metrajdaf3,
                        amar.metrajdaf4,
                        amar.metrajdaf5,
                        amar.metrajdaf6,
                        amar.metrajdaf7,
                        amar.metrajdaf8,
                    ]
                    # metraj_val_with_default = {key: value if value is not None else 0 for key, value in metraj_val.items()}
                    metraj_val = [value if value is not None else 0 for value in metraj_val]
                    saved_data_with_default = {key: value if value is not None else 0 for key, value in saved_data.items()}

                    makhraj_value = amar.makhraj_metraj_daf
                    if(not makhraj_value or makhraj_value==0):
                        makhraj_value=1

                    total_metraj = sum(value if value is not None else 0 for value in metraj_val)

                    # Calculate the sum of makhraj values
                    total_val = total_metraj / makhraj_value
                    machines_with_formulas.append({'machine': machine, 'formula': formula.formula,'speedformula':speedformula.formula,'amar':amar,'shift':s,'metraj':saved_data_with_default,'total_val':total_val})
                    # else:
                    #     machines_with_formulas.append({'machine': machine, 'formula': formula.formula,'speed':0,'nomre':0,'speedformula':speedformula.formula})


                except Formula.DoesNotExist:
                    machines_with_formulas.append({'machine': machine, 'formula': None,'formula': 0,'speed':0,'nomre':0})
                except SpeedFormula.DoesNotExist:
                    machines_with_formulas.append({'machine': machine, 'formula': None,'formula': 0,'speed':0,'nomre':0,'speedformula':0})
                except DailyProduction.DoesNotExist:
                    machines_with_formulas.append({'machine': machine, 'formula': formula.formula,'speed':0,'nomre':0,'speedformula':speedformula.formula})

        data['html_heatset_result'] = render_to_string('mrp/tolid/partialHeatsetList.html',{
            'machines':machines_with_formulas,
            'shifts':shift,'next_date':next_day.strftime('%Y-%m-%d'),'prev_date':previous_day.strftime('%Y-%m-%d'),'today':jdatetime.date.fromgregorian(date=date_object)}
        )
        data['prev_date']=previous_day.strftime('%Y-%m-%d')
        data['next_date']=next_day.strftime('%Y-%m-%d')
        data['today_shamsi']=str(jdatetime.date.fromgregorian(date=date_object))

        # return render(request,"mrp/tolid/daily_details.html",{'machines':machines_with_formulas,'shifts':shift,'next_date':next_day.strftime('%Y-%m-%d'),'prev_date':previous_day.strftime('%Y-%m-%d'),'today':jdatetime.date.fromgregorian(date=date_object),'title':'آمار روزانه'})
        return JsonResponse(data)
def list_amar_daily_info(request):

        data=dict()
        asset_category = AssetCategory.objects.all().order_by('priority')
        entry_forms = EntryForm.objects.prefetch_related('asset_details__asset_category')

    # Create a list of dictionaries in the desired format
        moshakhasat = []

        for entry in entry_forms:
            entry_data = entry
            asset_details = [
                {
                    "id": asset.id,
                    "asset_category": asset.asset_category.name,  # Assuming AssetCategory2 has a `name` field
                    "asset_category_id": asset.asset_category.id,  
                    "nomre": asset.nomre,
                    "speed": asset.speed,
                }
                for asset in entry.asset_details.all()
            ]
            try:
        
                moshakhasat.append({
                    "entry": entry_data,
                    "assetdetail": json.dumps(asset_details),
                })
            except:
                pass
        # annotate(
        # min_priority=models.Min('asset__assetTavali')
        # ).order_by('min_priority')

        dayOfIssue=request.GET.get('event',False)
        shift_id=request.GET.get('shift_id',False)
        
        if(not dayOfIssue):
            dayOfIssue=request.GET.get('event_id',datetime.datetime.now())
            date_object = DateJob.getTaskDate(dayOfIssue)
        else:
            date_object=datetime.datetime.strptime(str(dayOfIssue), '%Y-%m-%d')

        next_day = date_object + timedelta(days=1)

    # Calculate previous day
        previous_day = date_object - timedelta(days=1)
        machines=Asset.objects.filter(assetTypes=3)

        shift=Shift.objects.get(id=shift_id)
        machines_with_formulas = []
        # for s in shift:
        s=shift
        for machine in machines:
            try:

                formula = Formula.objects.get(machine=machine)
                speedformula = SpeedFormula.objects.get(machine=machine)
                amar=DailyProduction.objects.get(machine=machine,dayOfIssue=date_object,shift=s)

                machines_with_formulas.append({'machine': machine,'vahed':machine.assetVahed, 'formula': formula.formula,'speedformula':speedformula.formula,'amar':amar,'shift':s,'shift_id':s})

                # else:
                #     machines_with_formulas.append({'machine': machine, 'formula': formula.formula,'speed':0,'nomre':0,'speedformula':speedformula.formula})


            except DailyProduction.DoesNotExist:
                    # print("error",)
                    formula = Formula.objects.get(machine=machine)
                    speedformula = SpeedFormula.objects.get(machine=machine)
                    max_nomre = DailyProduction.objects.filter(machine=machine).aggregate(Max('nomre'))

                    new_daily_production = DailyProduction(
                    machine=machine,
                    zayeat=0,
                    shift=s,
                    dayOfIssue=dayOfIssue,
                    vahed=machine.assetVahed,
                    speed=0,
                    nomre=max_nomre['nomre__max'],
                    counter1=0,
                    counter2=0,

                    production_value=0,
                    daf_num=0,
                    dook_weight=0,
                    weight1=0,
                    weight2=0,
                    weight3=0,
                    weight4=0,
                    weight5=0,
                    net_weight=0,
                    metrajdaf1=0,
                    metrajdaf2=0,
                    metrajdaf3=0,
                    metrajdaf4=0,
                    metrajdaf5=0,
                    metrajdaf6=0,
                    metrajdaf7=0,
                    metrajdaf8=0,
                    makhraj_metraj_daf=1,
                    )


                    # Save the object to the database
                    # new_daily_production.save()

                    machines_with_formulas.append({'machine': machine, 'formula': None,'formula': formula.formula,'speedformula':speedformula.formula,'nomre':max_nomre['nomre__max'],'amar':new_daily_production,'shift':s,'speedformula':speedformula.formula,'vahed':machine.assetVahed,'shift_id':s})
            except Formula.DoesNotExist:
                machines_with_formulas.append({'machine': machine,'formula': 0,'speed':0,'nomre':0,'vahed':machine.assetVahed,'shift_id':s})
            except SpeedFormula.DoesNotExist:
                # amar=DailyProduction.objects.get(machine=machine,dayOfIssue=date_object,shift=s)
                machines_with_formulas.append({'machine': machine, 'formula': formula.formula,'speedformula':0 ,'formula': 0,'speed':0,'nomre':0,'speedformula':0,'vahed':machine.assetVahed,'shift_id':s})
            except DailyProduction.DoesNotExist:

                machines_with_formulas.append({'machine': machine, 'formula': formula.formula,'speed':0,'nomre':0,'speedformula':speedformula.formula,'vahed':machine.assetVahed,'shift_id':s})


        data['html_heatset_result'] = render_to_string('mrp/tolid/partialAssetAmarList.html',{
            'machines':machines_with_formulas,'cat_list':asset_category,'shift_id':s.id,'moshakhasat':moshakhasat,
            'shifts':shift,'next_date':next_day.strftime('%Y-%m-%d'),'prev_date':previous_day.strftime('%Y-%m-%d'),'today':jdatetime.date.fromgregorian(date=date_object)}
        )
        data['prev_date']=previous_day.strftime('%Y-%m-%d')
        data['next_date']=next_day.strftime('%Y-%m-%d')
        data['today_shamsi']=str(jdatetime.date.fromgregorian(date=date_object))

        # return render(request,"mrp/tolid/daily_details.html",{'machines':machines_with_formulas,'shifts':shift,'next_date':next_day.strftime('%Y-%m-%d'),'prev_date':previous_day.strftime('%Y-%m-%d'),'today':jdatetime.date.fromgregorian(date=date_object),'title':'آمار روزانه'})
        return JsonResponse(data)

#####################        heatset    ########################
def save_HeatsetMetraj_form(request, form, template_name):


    data = dict()
    if (request.method == 'POST'):
        if form.is_valid():

            # Handle form submission
            # Access form.cleaned_data to get the values
            # Save the data or perform any necessary actions
            saved_data = {
                'metrajdaf1': form.cleaned_data['metrajdaf1'],
                'metrajdaf2': form.cleaned_data['metrajdaf2'],
                'metrajdaf3': form.cleaned_data['metrajdaf3'],
                'metrajdaf4': form.cleaned_data['metrajdaf4'],
                'metrajdaf5': form.cleaned_data['metrajdaf5'],
                'metrajdaf6': form.cleaned_data['metrajdaf6'],
                'metrajdaf7': form.cleaned_data['metrajdaf7'],
                'metrajdaf8': form.cleaned_data['metrajdaf8'],
                'makhraj_metraj_daf': form.cleaned_data['makhraj_metraj_daf'],
            }
            metraj_values = [
                form.cleaned_data['metrajdaf1'],
                form.cleaned_data['metrajdaf2'],
                form.cleaned_data['metrajdaf3'],
                form.cleaned_data['metrajdaf4'],
                form.cleaned_data['metrajdaf5'],
                form.cleaned_data['metrajdaf6'],
                form.cleaned_data['metrajdaf7'],
                form.cleaned_data['metrajdaf8'],
            ]
            makhraj_value = form.cleaned_data['makhraj_metraj_daf']
            total_metraj = sum(metraj_values)

            # Calculate the sum of makhraj values
            total_val = total_metraj / makhraj_value

            # Here you can perform additional actions, such as saving the data to the database

            # Return a JSON response with the saved data
            return JsonResponse({'success': True, 'data': saved_data,'form_is_valid':True,'total_val': total_val})
        else:
            data['form_is_valid'] = False
            print(form.errors)

    context = {'form': form}


    data['html_heatsetmetraj_form'] = render_to_string(template_name, context, request=request)
    return JsonResponse(data)
def tolid_heatset_metraj_create(request):
    if (request.method == 'POST'):
        form = HeatsetMetrajForm(request.POST)
        return save_HeatsetMetraj_form(request, form, 'mrp/tolid/partialHeatsetMetrajCreate.html')
    else:
        data_is_ok=False
        initial_data=None
        try:
            metraj_data=eval(json.loads(request.GET.get("data",False)))

            initial_data=metraj_data
            data_is_ok=True

        except  Exception as ex:
            # print(request.GET.get("data",False))
            print(ex)
            pass

        if(data_is_ok==False):
            initial_data = {'metrajdaf1': 0, 'metrajdaf2': 0, 'metrajdaf3': 0, 'metrajdaf4': 0,
                        'metrajdaf5': 0, 'metrajdaf6': 0, 'metrajdaf7': 0, 'metrajdaf8': 0,
                        'makhraj_metraj_daf': 1}
        form = HeatsetMetrajForm(initial=initial_data)



        return save_HeatsetMetraj_form(request, form, 'mrp/tolid/partialHeatsetMetrajCreate.html')
def delete_heatset_info(request):
    dayOfIssue=request.GET.get('event_id',False)
    date=DateJob.getTaskDate(request.GET.get('event_id',False))
    heatset_amar=DailyProduction.objects.filter(dayOfIssue=dayOfIssue,mechine__assetCategory=8)
    shift=Shift.objects.all()
    for i in heatset_amar:
        i.delete()
    data=dict()
    data['html_heatset_result'] = render_to_string('mrp/tolid/partialHeatsetList.html',{
        'machines':'',
        'shifts':shift,'next_date':'','prev_date':'','today':''}
    )
    return JsonResponse(data)
def delete_heatset_info(request):
    dayOfIssue=request.GET.get('event_id',False)
    date=DateJob.getTaskDate(request.GET.get('event_id',False))
    heatset_amar=DailyProduction.objects.filter(dayOfIssue=dayOfIssue,mechine__assetCategory=8)
    shift=Shift.objects.all()
    for i in heatset_amar:
        i.delete()
    data=dict()
    data['html_heatset_result'] = render_to_string('mrp/tolid/partialHeatsetList.html',{
        'machines':'',
        'shifts':shift,'next_date':'','prev_date':'','today':''}
    )
    return JsonResponse(data)
def delete_amar_info(request):
    dayOfIssue=request.GET.get('event_id',False)
    date=DateJob.getTaskDate(request.GET.get('event_id',False))
    heatset_amar=DailyProduction.objects.filter(dayOfIssue=dayOfIssue,machine__assetTypes=2)
    shift=Shift.objects.all()
    print(heatset_amar.count())
    for i in heatset_amar:
        i.delete()
    data=dict()
    data['html_heatset_result'] = render_to_string('mrp/tolid/partialAssetAmarList.html',{
        'machines':'',
        'shifts':shift,'next_date':'','prev_date':'','today':''}
    )
    return JsonResponse(data)
