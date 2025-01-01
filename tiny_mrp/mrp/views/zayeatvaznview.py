from django.shortcuts import render
from mrp.models import *
import jdatetime
from django.db import IntegrityError
from django.views.decorators.csrf import csrf_exempt
import json
from django.http import JsonResponse
from mrp.business.DateJob import *
from datetime import datetime, timedelta
from django.contrib.contenttypes.models import ContentType
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.context_processors import PermWrapper
from django.contrib.auth.decorators import login_required
from django.template.loader import render_to_string
from django.views.decorators import csrf
from mrp.business.tolid_util import *
from django.http import HttpResponse
from openpyxl import Workbook
import pickle
from collections import defaultdict

@login_required
@csrf_exempt
def zayeatVazn_create(request):

    if (request.method == 'POST'):
        try:
            # Assuming the data is sent as JSON
            received_data = json.loads(request.body)  # If data was sent as form-encoded, use request.POST
            # Process the received_data (In this case, it's assumed to be a list of dictionaries)
            print(received_data)
            for table in received_data:


                    for row in table:
                        
                        ff=ZayeatVaz.objects.filter(zayeat=Zayeat.objects.get(id=row['id']),shift=Shift.objects.get(id=row['shift']),dayOfIssue=row['date'],makan=Asset.objects.get(id=row['makan']))
                        if(ff.count()>0):
                            z=ff[0]
                            z.vazn=float(row['vazn'])
                            z.zayeat=Zayeat.objects.get(id=row['id'])
                            z.dayOfIssue=row['date']
                            z.shift=Shift.objects.get(id=row['shift'])
                            z.save()

                        else:
                            z=ZayeatVaz()
                            z.vazn=float(row['vazn'])
                            z.zayeat=Zayeat.objects.get(id=row['id'])
                            z.dayOfIssue=row['date']
                            z.shift=Shift.objects.get(id=row['shift'])
                            z.makan=Asset.objects.get(id=row['makan'])
                            z.save()

            # For demonstration purposes, just returning the received data as JSON response
            return JsonResponse({'success': True, 'data_received': received_data})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'error': str(e)})
    else:
        data=dict()
        date_of_issue=None

        current_date=request.GET.get("data",False)
        current_makan=request.GET.get("makan",False)
        if(current_date):
            date_of_issue=DateJob.getTaskDate(current_date)
        else:
            date_of_issue=datetime.now().date()
        za=Zayeat.objects.all()
        date_zayeat=ZayeatVaz.objects.filter(dayOfIssue=date_of_issue,makan__id=current_makan)
        shift=Shift.objects.all()
        zayeat_vazn_dict = defaultdict(list)
        for zv in date_zayeat:
            zayeat_vazn_dict[zv.zayeat.id].append({'vazn':zv.vazn,'shift':zv.shift.id})
        data['data']=render_to_string('mrp/zayeat_vazn/partialZayeatVaznCreate.html',
            {   'shifts':shift,
                'zayeat':za,
                'zayeat_vazn':zayeat_vazn_dict,
                'makan':current_makan,
                'date':date_of_issue.strftime('%Y-%m-%d')
            },request
        )
        return JsonResponse(data)
def get_daily_zaye(request):
    dayOfIssue=request.GET.get('event_id',datetime.now())
    makan_id=request.GET.get("makan_id",False)
    makan_name=Asset.objects.get(id=makan_id).assetName
    date_object = datetime.strptime(dayOfIssue, '%Y-%m-%d')
    za=Zayeat.objects.all()
    date_zayeat=ZayeatVaz.objects.filter(dayOfIssue=date_object,makan__id=makan_id)
    shift=Shift.objects.all()
    zayeat_vazn_dict = defaultdict(list)
    for zv in date_zayeat:
        zayeat_vazn_dict[zv.zayeat.id].append({'vazn':round(zv.vazn, 2),'shift':zv.shift.id})
    return render(request,'mrp/zayeat_vazn/zayeatVaznList.html',
        {   'shifts':shift,
            'makan_name':makan_name,
            'makan_id':makan_id,
            'zayeat':za,
            'zayeat_vazn':zayeat_vazn_dict,
            'date':date_object,'jalali':jdatetime.date.fromgregorian(date=date_object).strftime('%d-%m-%Y')
        }
    )
def monthly_zayeat_detaild_report(request):
    makan_id=request.GET.get("makan_id",False)
    makan=Asset.objects.filter(assetIsLocatedAt__isnull=True)
    days=[]
    shift=Shift.objects.all()
    saloons=Asset.objects.filter(assetIsLocatedAt__isnull=True)
    salooon_shifts=[]
    for saloon in saloons:
        for sh in shift:
            salooon_shifts.append({'saloon_id':saloon.id,'saloon_name':saloon.assetName,'shift_name':sh.name,'shift_id':sh.id})


    # asset_category = AssetCategory.objects.all().order_by('priority')

    current_date_time2 = jdatetime.datetime.now()
    current_year=current_date_time2.year
    j_month=request.GET.get('month',current_date_time2.month)

    j_year=int(request.GET.get('year',current_year))
    current_date_time = jdatetime.date(j_year, int(j_month), 1)
    current_jalali_date = current_date_time
    if current_jalali_date.month == 12:
        first_day_of_next_month = current_jalali_date.replace(day=1, month=1, year=j_year + 1)
    else:
        first_day_of_next_month = current_jalali_date.replace(day=1, month=current_jalali_date.month + 1)


    num_days = (first_day_of_next_month - jdatetime.timedelta(days=1)).day
    cat_list=[]
    days=[]
    counter=0
    sum_kol=0
    sum_dic={}
    for saloon in saloons:
            sum_dic[str(saloon.id)]={}
            for sh in shift:
                sum_dic[str(saloon.id)][str(sh.id)]=0

    for day in range(1,num_days+1):
        product=[]
        sum=0

        for saloon in saloons:         
          

        
            j_date=jdatetime.date(j_year,current_jalali_date.month,day)
            # print(j_date,'!!!!!!!!!!!!!')
            for sh in shift:
                z_val=get_sum_zayeat_by_date_shift_makan(saloon.id,sh,j_date.togregorian())
                sum+=z_val["total_vazn"] if z_val["total_vazn"]  else 0
                tmp=z_val["total_vazn"] if z_val["total_vazn"]  else 0
                sum_dic[str(saloon.id)][str(sh.id)]+=tmp
                # a_tmp=sum_dic[str(saloon.id)][str(sh.id)]
                # print("tmp",tmp)
                product.append({'shift':sh,'value':z_val})#get_sum_machine_by_date_shift_makan(cats,makan_id,sh,j_date.togregorian())
                counter+=1
                # print("!!!!!!!!!",cats,product[sh.id])
        days.append({'saloon':saloon,'sum':sum,'date':"{0}/{1}/{2}".format(j_year,current_jalali_date.month,day),'day_of_week':DateJob.get_day_of_week(j_date),'product':product})
        sum_kol+=sum
        # product={}
        start=jdatetime.date(j_year,current_jalali_date.month,1)
        end=jdatetime.date(j_year,current_jalali_date.month,num_days)
    tt=[]
    ts=[]
    for i,l in sum_dic.items():
            sum_s=0
            for m,n in l.items():
                sum_s+=n
                tt.append({"saloon":i,"shift":m,"val":n})
            ts.append({"saloon":i,"sum_saloon":sum_s})
            
            
    print(sum_dic)
    return render(request,'mrp\zayeat_vazn\monthly_zayeat.html',{'sum_s':ts,'sum_shif':tt,'sum':sum_kol,'makan':makan,'cats':days,'title':'آمار ماهانه ضایعات','cat_list':cat_list,'shift':salooon_shifts,'month':j_month,'year':j_year,'saloon':saloons})

def export_monthly_zayeat(request):
    # Create an in-memory workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data Export"
    makan=Asset.objects.filter(assetIsLocatedAt__isnull=True)
    days=[]
    shift=Shift.objects.all()
    saloons=Asset.objects.filter(assetIsLocatedAt__isnull=True)
    salooon_shifts=[]
    # Add two rows of headers
    makan_names = [m.assetName for m in makan]
    print(makan_names)
    sheet.append(["ID"] + makan_names) 
    sheet.append(["", "A", "B", "C", "A", "B", "C", "A", "B", "C"])

    # Merge cells for grouped headers
    # sheet.merge_cells("A1:C1")  # Merge for "Header Group 1"
    # sheet.merge_cells("D1:E1")  # Merge for "Header Group 2"

    # Add your data
    data = [
        [1, "Asset 1"] + ["Available"] * len(makan_names),
        [2, "Asset 2"] + ["Not Available"] * len(makan_names),
    ]
    for row in data:
        sheet.append(row)

    # Prepare the HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=data_export.xlsx"

    # Save the workbook to the response
    workbook.save(response)

    return response